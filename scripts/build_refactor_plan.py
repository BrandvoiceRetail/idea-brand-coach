#!/usr/bin/env python3
"""Generate the Refactor & Architecture Build Plan workbook + markdown report.

BMad Master deliverable (2026-05-26). FOCUS: v2-and-beyond.
Steering constraints from Matthew:
  - v2 is canonical; v1 is legacy (/v1/*) and stays in place.
  - Do NOT delete anything yet.
  - Do NOT refactor old (v1) code unless it supports current/planned functionality.
Grounded in verified code findings + graphify + both sprint plans + Launch Roadmap v2.
"""
from __future__ import annotations
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

OUT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "_bmad-output")
XLSX = os.path.join(OUT_DIR, "refactor_build_plan_v1_2026-05-26.xlsx")
MD = os.path.join(OUT_DIR, "refactor_build_plan_v1_2026-05-26.md")

H1 = Font(bold=True, size=14, color="1F2937")
H2 = Font(bold=True, size=11, color="FFFFFF")
BOLD = Font(bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")
TOP = Alignment(vertical="top")
HDR_FILL = PatternFill("solid", fgColor="374151")
PRI_FILL = {
    "P0": PatternFill("solid", fgColor="FCA5A5"),
    "P1": PatternFill("solid", fgColor="FDBA74"),
    "P2": PatternFill("solid", fgColor="FDE68A"),
    "P3": PatternFill("solid", fgColor="BBF7D0"),
    "P4": PatternFill("solid", fgColor="E5E7EB"),
}
KEEP_FILL = PatternFill("solid", fgColor="BBF7D0")
FREEZE_FILL = PatternFill("solid", fgColor="FDE68A")
DEFER_FILL = PatternFill("solid", fgColor="E5E7EB")
THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def hdr(ws, row, headers, widths=None):
    for c, text in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=text)
        cell.font = H2; cell.fill = HDR_FILL; cell.alignment = WRAP; cell.border = BORDER
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i)].width = w


def put(ws, row, values, wrap_cols=None, fill=None):
    wrap_cols = wrap_cols or set()
    for c, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.alignment = WRAP if c in wrap_cols else TOP
        cell.border = BORDER
        if fill:
            cell.fill = fill


def kv_sheet(ws, title, pairs, kw=28, vw=110):
    ws.column_dimensions["A"].width = kw
    ws.column_dimensions["B"].width = vw
    ws["A1"] = title; ws["A1"].font = H1; ws.merge_cells("A1:B1")
    r = 3
    for k, v in pairs:
        ws.cell(row=r, column=1, value=k).font = BOLD
        ws.cell(row=r, column=1).alignment = WRAP
        ws.cell(row=r, column=2, value=v).alignment = WRAP
        r += 1


wb = Workbook()

# === SHEET 1 — Overview =====================================================
ws = wb.active; ws.title = "Overview"
kv_sheet(ws, "IDEA Brand Coach — Refactor & Architecture Build Plan (v2-and-beyond)", [
    ("Version / Date", "v1 — 2026-05-26. Author: BMad Master. Scope: v2 experience + shared edge-function backend."),
    ("Objective", "Make v2 a maintainable, viable app that gives users the best UX solving their problem, where technical errors are "
        "identified, logged, and fixed FAST — sequenced to serve the Alpha->Beta->GA roadmap."),
    ("STEERING (Matthew, 2026-05-26)", "1) v2 is canonical. 2) v1 (/v1/* legacy namespace) stays in place — DO NOT DELETE ANYTHING YET. "
        "3) Do NOT refactor old code unless it supports current or planned functionality. This plan touches v2 + shared backend ONLY."),
    ("Governing constraint", "Refactor spend <=20-30% of any 20h sprint (<=4-6h), unless an item is massive-impact (justified inline). "
        "Current ship week (Sprint v3) is REFACTOR-FROZEN. Work below schedules into the next sprints' buffer / Always-On."),
    ("Method / honesty", "Every finding re-verified against real code (grep + file reads), not graph inference alone. Routing confirmed "
        "v1 is fully namespaced under /v1/* with VersionGate -> /v2/coach as the live entry. Unverified items marked AMBIGUOUS."),
    ("The 4 product primitives", "Nearly every roadmap feature reduces to: (1) conversation -> structured fields; (2) local-first edit-aware "
        "persistence; (3) Trevor-voiced LLM artifact; (4) feedback moment. The target abstractions ARE these — built once, in v2/shared."),
    ("v1 disposition (summary)", "KEEP (load-bearing now): /free-diagnostic, /beta*, /v1/canvas. FREEZE: the rest of /v1/* — leave untouched, "
        "no refactor, no delete. FUTURE: decommission is a real LOC win but DEFERRED until v2 supersedes the kept surfaces. See 'V1 Disposition' tab."),
    ("Headline recommendation", "Spend the first reclaimed budget on the OBSERVABILITY SEAM (RF-01/RF-02). It is simultaneously an Alpha roadmap "
        "GATE, the user's #1 goal, and risk-reduction for 12+ shared LLM edge functions. Treat as the massive-impact exception."),
    ("How to use", "1 Overview | 2 Target Abstractions | 3 Refactor Backlog (priority-ordered; update Status) | 4 Observability Plan | "
        "5 Sequencing (<=30%/sprint) | 6 V1 Disposition | 7 Out of Scope."),
])

# === SHEET 2 — Target Abstractions ==========================================
ws = wb.create_sheet("Target Abstractions")
hdr(ws, 1, ["#", "Abstraction (v2/shared)", "Verdict", "Current state (verified)", "What to build / change", "Rationale & roadmap leverage"],
    widths=[4, 26, 16, 46, 46, 50])
abstractions = [
    ("1", "Single Field-Extraction pipeline (v2)", "ACCEPT",
     "Within v2/shared the conversation->fields logic appears in 3 places: v2/components/EnhancedChatInterface.tsx (extractFieldsFromContent), "
     "inline regex in V2ChatPanel, and hooks/useFieldExtraction.ts (474L) used by v2. (v1 utils/fieldExtractionParser.ts is LEFT ALONE.)",
     "One canonical v2 extractor, schema-driven off CHAPTER_FIELDS_MAP; v2 callers delegate to it. Do not touch v1's copy.",
     "Core v2 flow. Unblocks Avatar 2.0 (F-004..007), Review Lab (F-063), Decision Trigger (F-058)."),
    ("2", "One v2 Field Repository (edit-aware)", "ACCEPT",
     "v2 sync hooks (useAvatarFieldSync, useFieldDatabaseSync, useSimpleFieldSync, useFieldSync) sit over the knowledge-base spine "
     "(knowledge-repository.ts 344L + supabase-sync-service.ts 429L, already implements I*Repository/ISyncService w/ retry+backoff). "
     "v1 lib/sync/FieldSyncService.ts + services/field/FieldValueStorageService.ts serve v1 only.",
     "Converge the v2 field hooks onto the knowledge-base repository; manual-edit lock in EXACTLY one place. Freeze v1 sync code in place.",
     "Trust foundation. Alpha gate + QA require 'manual edits never silently overwritten'. Highest correctness risk."),
    ("3", "Typed LLM edge-function client (shared)", "ACCEPT",
     "23 edge functions, ~12 LLM-calling (brand-ai-assistant, reveal-signature, diagnostic-interpretation, brand-copy-generator, "
     "ai-insight-guidance, contextual-help, buyer-intent-analyzer, generate-brand-strategy-section...). No shared typed client; quota "
     "errors hide in HTTP-200 SSE bodies (memory: feedback_sse_stream_capture). Backend is shared by v1 AND v2.",
     "One typed client: invoke + SSE parse + quota/error-envelope detection + retry/timeout. All current + future generators are instances.",
     "The Create-mode generator base (F-020/021/055/063). Closes the riskiest observability gap on the most failure-prone surface."),
    ("4", "Result<T> propagated + reportError sink", "ACCEPT (cheap)",
     "ErrorBoundary EXISTS (components/ErrorBoundary.tsx 122L; shared). Result<T> EXISTS in v2 (v2/shared/types, domain, "
     "lib/knowledge-base/interfaces). NO reportError/captureException/Sentry anywhere.",
     "Add one reportError() sink (console+overlay now, Sentry-ready interface). Wire ErrorBoundary to it. Propagate Result<T> across v2 boundaries.",
     "Satisfies Alpha 'error monitoring' gate cheaply. Identify->log->fix loop. Most infra already seeded in v2 — wiring, not building."),
    ("5", "Enforce DI data access (v2/shared)", "ACCEPT (enforce, don't build)",
     "DI spine holds: ServiceProvider + interface-typed services. Real leaks in scope: ~2 files in hooks/v2 import the raw client. "
     "(v1 chat-component leaks are LEFT ALONE per steering.)",
     "Retire the hooks/v2 raw-client imports onto useServices()/interfaces. Add an eslint guard scoped to v2.",
     "Decouples the supabase god-node (57 edges) for v2; enables clean mocking. Incremental / boy-scout."),
    ("6", "FeedbackMoment primitive (shared)", "ACCEPT",
     "Write path exists (BetaFeedbackWidget.tsx -> feedback_events). Roadmap defines 3 moments used across ALL phases; this week ships Moment 1.",
     "Extract <FeedbackMoment moment={1|2|3}> writing feedback_events with consistent payload + skip-as-data behaviour. Used by beta flow + v2.",
     "Moments 1/2/3 recur Alpha->Beta->GA (Feedback Loop sheet). Build once; Branch-A next sprint adds Moment 2."),
    ("7", "v1-vs-v2 canonical", "RESOLVED -> v2",
     "Routing: whole v1 surface namespaced /v1/* as 'legacy support'; VersionGate -> /v2/coach is the live entry. useBrandCoachV2State is "
     "mid-extraction into useChatOrchestration/useFieldOrchestration/useBrandCoachUI (worst-of-both right now).",
     "No v1 deletion now. In-scope action is COMPLETING the v2 orchestration extraction (RF-07 below) — that is current live functionality.",
     "Resolves the risk-register 'two warm codebases' item by direction (v2 wins) without touching v1."),
]
r = 2
for row in abstractions:
    put(ws, r, row, wrap_cols={2, 4, 5, 6}); r += 1
ws.freeze_panes = "A2"

# === SHEET 3 — Refactor Backlog =============================================
ws = wb.create_sheet("Refactor Backlog")
cols = ["ID", "Priority", "Refactor (v2/shared)", "Why", "Key files", "Effort (h)", "Risk", "Blast radius",
        "Leverage", "Roadmap F-IDs", "Recommended sprint", "Allocation", "Status"]
hdr(ws, 1, cols, widths=[8, 9, 26, 38, 42, 11, 8, 12, 10, 18, 18, 20, 12])
backlog = [
    ("RF-01", "P0", "Typed LLM edge-fn client (abstraction #3)",
     "12+ shared LLM edge fns each hand-roll invoke/SSE/error handling; quota errors hide in HTTP-200 bodies -> silent failures users hit.",
     "NEW src/services/llm/EdgeFunctionClient.ts; supabase/functions/_shared; wraps brand-ai-assistant, reveal-signature, diagnostic-interpretation +9",
     "3-4 (thin)/6-8 (full)", "Med", "High (12+ fns)", "High",
     "F-020/021/055/063", "Next sprint (06-01)", "Thin slice 3-4h", "not_started"),
    ("RF-02", "P0", "reportError sink + Result<T> propagation (abstraction #4)",
     "No error reporting sink; Alpha gate REQUIRES 'see errors before testers report them'. ErrorBoundary + Result<T>(v2) already exist.",
     "src/components/ErrorBoundary.tsx; NEW src/lib/observability/reportError.ts; src/v2/shared/types/index.ts",
     "2-3", "Low", "Med (v2)", "High",
     "Alpha error GATE", "Next sprint (06-01)", "2-3h", "not_started"),
    ("RF-03", "P1", "Converge v2 Field Repository + single edit-lock (abstraction #2)",
     "v2 field-sync hooks duplicate persistence over the knowledge-base spine; manual-edit lock (the trust foundation) is spread out.",
     "src/hooks/{useAvatarFieldSync,useFieldDatabaseSync,useSimpleFieldSync,useFieldSync}.ts -> src/lib/knowledge-base/* (KEEP target)",
     "6-8 (split 2 sprints)", "HIGH", "High (v2 fields)", "High",
     "Alpha 'no overwrite' GATE", "Sprints +2 / +3", "4h then 3h", "not_started"),
    ("RF-04", "P1", "Finish useBrandCoachV2State extraction (abstraction #7 action)",
     "Live /v2/coach state hook is mid-extraction into 3 orchestrators that duplicate slices = worst of both. This is CURRENT functionality.",
     "src/hooks/v2/{useBrandCoachV2State,useChatOrchestration,useFieldOrchestration,useBrandCoachUI}.ts",
     "3-5", "Med", "High (live coach)", "High",
     "Coach mode (F-022)", "Sprint +2", "3-5h", "not_started"),
    ("RF-05", "P2", "Unify v2 Field-Extraction pipeline (abstraction #1)",
     "3 reimplementations of conversation->fields inside v2 drift apart; every extracting feature multiplies divergence.",
     "src/v2/components/EnhancedChatInterface.tsx; V2ChatPanel inline regex; src/hooks/useFieldExtraction.ts (v1 copy left alone)",
     "4-6", "Med", "High", "High",
     "F-004..007, F-058, F-063", "Sprint +3", "4-6h", "not_started"),
    ("RF-06", "P2", "Extract FeedbackMoment primitive (abstraction #6)",
     "Moments 1/2/3 recur across all phases; building each inline duplicates write/skip/payload logic.",
     "src/components/BetaFeedbackWidget.tsx; src/pages/BetaFeedback.tsx; NEW src/components/feedback/FeedbackMoment.tsx",
     "2-3", "Low", "Med", "Med",
     "Feedback Moments 1/2/3", "Branch-A next sprint", "2-3h", "not_started"),
    ("RF-07", "P3", "Enforce DI data access in v2 (abstraction #5)",
     "~2 hooks/v2 files leak the raw supabase client, coupling v2 UI to the data plane and blocking mocking.",
     "src/hooks/v2/* (raw client imports); eslint no-restricted-imports guard scoped to v2 (v1 excluded)",
     "1-2 (incremental)", "Low", "Low (v2)", "Med",
     "Testability", "Boy-scout / Always-On", "~0.5h per touch", "not_started"),
    ("RF-08", "P4", "Collapse v2 cosmetic duplicates",
     "Parallel implementations of the same v2 UX: hygiene only, but inflate surface area and confuse contributors.",
     "v2: MilestoneCelebration vs MilestoneOverlay; GhostTextChatInput vs GhostTextFieldWrapper; useRejectionMessages vs useRejectionToChat; "
     "BrandsList vs SimpleBrandPanel; shared sonner vs toast/toaster",
     "3-4 (spread)", "Low", "Low", "Low",
     "(hygiene)", "Always-On / buffer", "opportunistic", "not_started"),
]
r = 2
for row in backlog:
    put(ws, r, row, wrap_cols={3, 4, 5, 9, 10, 11, 12})
    ws.cell(row=r, column=2).fill = PRI_FILL[row[1]]; ws.cell(row=r, column=2).font = BOLD
    r += 1
dv = DataValidation(type="list", formula1='"not_started,scheduled,in_progress,blocked,complete"', allow_blank=True)
ws.add_data_validation(dv); dv.add(f"M2:M{r-1}")
ws.freeze_panes = "A2"
ws.cell(row=r + 1, column=1, value="TOTAL (v2/shared)").font = BOLD
ws.cell(row=r + 1, column=3, value="~24-35h across 3+ sprints @ <=6h/sprint. v1 NOT included (frozen). RF-01/02 first.")

# === SHEET 4 — Observability Plan ===========================================
ws = wb.create_sheet("Observability Plan")
kv_sheet(ws, "Observability Plan — identify -> log -> fix, fast (v2 + shared backend)", [
    ("Goal", "A technical error is (a) IDENTIFIED before a tester reports it, (b) LOGGED with reproduce-able context, (c) FIXED fast. "
        "This is the Alpha 'Error monitoring' gate and the user's #1 objective."),
    ("Layer 1 — Edge-fn envelope", "An LLM edge fn can return HTTP 200 while the SSE body carries a billing/quota/length error (memory). "
        "RF-01: the typed client inspects the stream, classifies (quota|rate-limit|content|network|unknown), returns a typed error not a silent blank."),
    ("Layer 2 — Result<T> boundaries", "Every v2 service/edge call returns Result<T> (already the v2 convention). UI branches ok/err and renders "
        "a real error state instead of a blank/partial screen. RF-02 propagates it across v2 call sites."),
    ("Layer 3 — reportError sink", "Single src/lib/observability/reportError.ts(error, context): structured console + a ring buffer in a dev overlay. "
        "Sentry-ready interface (swap transport, not call sites). RF-02."),
    ("Layer 4 — ErrorBoundary wiring", "Existing ErrorBoundary calls reportError and shows a recoverable fallback. Add route-level boundaries around "
        "the diagnostic, signature, and /v2/coach flows so one feature crash never blanks the app."),
    ("Layer 5 — Supabase reachability", "Free-tier auto-pause -> NXDOMAIN/timeouts/INACTIVE (memory: project_supabase_pauses). Add a boot health ping "
        "+ a user-facing 'service paused' banner instead of a silent failure."),
    ("Sentry-ready seam", "Do NOT install Sentry this cycle (T10/F-017 deferred). Build the seam so adopting Sentry at Beta is a one-file transport swap. "
        "Console + dev overlay suffices for the alpha friend-tester cohort."),
    ("Definition of done", "Walk diagnostic->signature->feedback with Supabase PAUSED and with a forced edge-fn quota error. Both yield a clear recoverable "
        "error AND reportError captures cause + route + payload shape."),
])

# === SHEET 5 — Sequencing ===================================================
ws = wb.create_sheet("Sequencing")
hdr(ws, 1, ["Sprint / window", "Mode", "Refactor budget (<=30%)", "Scheduled items", "Dependency / rationale"],
    widths=[26, 22, 22, 40, 56])
seq = [
    ("Sprint v3 — wk 2026-05-25", "SHIP (frozen)", "0h (frozen)", "(none)",
     "Ship-week: Trust Gap + Signature on critical path T8->T3->T4->T5->T9. Refactoring here = scope creep per the plan."),
    ("Sprint +1 — wk 2026-06-01", "Conditional A/B/C", "<=6h (30%)", "RF-02 (2-3h) + RF-01 thin slice (3-4h)",
     "Observability first: it's an Alpha gate AND helps Branch-B diagnose WHY a v2 output failed. Confirm v2-canonical direction in planning (no v1 action)."),
    ("Sprint +2 — wk 2026-06-08", "Build", "<=6h (30%)", "RF-04 finish v2 coach extraction (3-5h) + RF-07 incremental (1h)",
     "Stabilise live /v2/coach before deeper field work. RF-04 is current functionality, so high value within budget."),
    ("Sprint +3 — wk 2026-06-15", "Build", "<=6h (30%)", "RF-03 pt1 (4h) + RF-06 (2h)",
     "Field Repository convergence depends on the observability seam (debuggable migration). FeedbackMoment slots into Branch-A feature work."),
    ("Sprint +4 — wk 2026-06-22", "Build", "<=6h (30%)", "RF-03 pt2 (3h) + RF-05 (3h)",
     "Finish repository convergence before unifying extraction (extraction writes through the repo)."),
    ("Always-On / buffer", "Continuous", "opportunistic", "RF-07 DI touches; RF-08 cosmetic dupes",
     "Hygiene with no gate dependency. Touch when already in the file (boy-scout), never as standalone sprint cost."),
    ("Massive-impact exception", "Note", "—", "RF-01 + RF-02 (~5-7h)",
     "Exceed the 'cheap' bar but justified: simultaneously an Alpha GATE, the #1 goal, and risk-reduction for 12+ edge fns. If Sprint +1 is "
     "Branch-B, they are arguably ON the critical path, not refactor overhead."),
]
r = 2
for row in seq:
    put(ws, r, row, wrap_cols={4, 5}); r += 1
ws.freeze_panes = "A2"

# === SHEET 6 — V1 Disposition ===============================================
ws = wb.create_sheet("V1 Disposition")
ws["A1"] = "V1 Disposition — NO ACTION NOW (no delete, no refactor)"; ws["A1"].font = H1; ws.merge_cells("A1:E1")
ws["A2"] = ("Per steering: v1 stays in place. This tab records the disposition so a future decommission is a deliberate, evidence-backed "
            "decision — NOT something done now. Classification by route reachability from the live (non-/v1) flow.")
ws["A2"].alignment = WRAP; ws.merge_cells("A2:E2")
hdr(ws, 4, ["v1 surface", "Disposition", "Evidence / reachability", "Why", "Future action (deferred)"],
    widths=[34, 16, 40, 40, 38])
v1rows = [
    ("/free-diagnostic + DiagnosticResults", "KEEP", "Live flow step 3-5 (sprint User Flow Contract); not under VersionGate redirect.",
     "Feeds the Trust Gap Score; current + planned functionality depends on it.",
     "None. Eventually re-skin under v2 shell."),
    ("/beta, /beta-journey, /beta-feedback", "KEEP", "Live alpha onboarding + Moment-1 feedback (T5/T6/T7 this sprint).",
     "Active tester funnel; FeedbackMoment primitive (RF-06) builds on it.",
     "None. Fold into v2 onboarding at Beta."),
    ("/v1/canvas (Brand Canvas)", "KEEP", "Sprint Fri-QA: 'Existing /brand-canvas still works — no regressions.'",
     "Still referenced as a working surface; Signature->Canvas bridge is planned (Branch-A T-A2).",
     "Re-implement as v2 Canvas when the bridge ships; then retire v1 canvas."),
    ("/v1/idea/* (insight, distinctive, empathy, authenticity, consultant)", "FREEZE", "Reachable only via /v1/* legacy redirects; not linked from VersionGate or /v2/coach.",
     "Superseded by the v2 coach experience. Not current/planned functionality.",
     "Decommission candidate AFTER v2 coach fully covers the IDEA modules."),
    ("/v1/start-here, /v1/journey, /v1/dashboard", "FREEZE", "Legacy entry/nav; live entry is VersionGate -> /v2/coach.",
     "Replaced by v2 entry. No current dependency.", "Decommission candidate."),
    ("/v1/avatar, /v1/brand-diagnostic, /v1/idea-diagnostic", "FREEZE", "Legacy /v1/* only.",
     "Avatar 2.0 ships in v2; v1 avatar/diagnostics superseded.", "Decommission candidate."),
    ("/v1/copy-generator, /v1/research-learning, /v1/conversations", "FREEZE", "Legacy /v1/* only.",
     "Create-mode + research re-built in v2/Beta scope.", "Decommission candidate."),
    ("v1-only services/hooks (FieldSyncService, FieldValueStorageService, fieldExtractionParser)", "FREEZE", "Imported by v1 surfaces; not the v2 knowledge-base spine.",
     "Leaving them avoids touching old code that doesn't serve current objectives.",
     "Remove with their v1 surfaces during decommission."),
]
r = 5
for row in v1rows:
    disp = row[1]
    fill = KEEP_FILL if disp == "KEEP" else (FREEZE_FILL if disp == "FREEZE" else DEFER_FILL)
    put(ws, r, row, wrap_cols={1, 3, 4, 5})
    ws.cell(row=r, column=2).fill = fill; ws.cell(row=r, column=2).font = BOLD
    r += 1
ws.cell(row=r + 1, column=1, value="Decommission trigger (future)").font = BOLD
ws.cell(row=r + 1, column=3, value="Only when v2 fully covers the KEEP surfaces (diagnostic, beta, canvas) AND analytics show /v1/* traffic ~0. "
                                   "Then do a measured, test-guarded removal — its own task, never inside a ship week.").alignment = WRAP
ws.merge_cells(start_row=r + 1, start_column=3, end_row=r + 1, end_column=5)
ws.freeze_panes = "A5"

# === SHEET 7 — Out of Scope =================================================
ws = wb.create_sheet("Out of Scope")
hdr(ws, 1, ["Item", "Why out of scope (now)", "Revisit when"], widths=[40, 70, 30])
oos = [
    ("Deleting ANY v1 code", "Explicit steering: no deletion yet. v1 stays in place; disposition recorded on the V1 Disposition tab.",
     "Future decommission task once v2 supersedes the KEEP surfaces."),
    ("Refactoring v1 internals", "Steering: do not touch old code unless it serves current/planned functionality. The KEEP surfaces work as-is.",
     "Only if a KEEP surface (diagnostic/beta/canvas) needs a change for a planned feature."),
    ("Full Sentry / APM install (F-017)", "Console + dev overlay + reportError seam is enough for the alpha cohort. Installing now is cost without payoff.",
     "Beta (paying users) — swap the reportError transport."),
    ("Rewriting the 23 edge functions", "RF-01 wraps them behind one client; rewriting internals is unnecessary and high-risk.",
     "Only if a specific fn proves unreliable in alpha."),
    ("Test-coverage backfill", "Separate from structural refactor. Add characterization tests AROUND each refactor as you make it.",
     "Per-refactor, inline."),
    ("Design-system / shadcn cleanup", "cn()/Button/Card god-nodes are unavoidable shared primitives. Toast dedupe is the only real item (RF-08).",
     "Opportunistic only."),
    ("Performance optimization", "No evidence of a perf problem in the graph or transcripts. Premature.",
     "When a measured perf issue appears."),
]
r = 2
for row in oos:
    put(ws, r, row, wrap_cols={1, 2, 3}); r += 1
ws.freeze_panes = "A2"

os.makedirs(OUT_DIR, exist_ok=True)
wb.save(XLSX)
print("workbook ->", XLSX)

# === Markdown report ========================================================
md = """# IDEA Brand Coach — Refactor & Architecture Build Plan (v1, 2026-05-26)
### Focus: **v2 and beyond** · v1 stays in place (no delete, no refactor)

**Author:** BMad Master · **Method:** graphify candidates re-verified against real code (grep + file reads); unverified items marked AMBIGUOUS.

## Steering constraints (Matthew, 2026-05-26)
1. **v2 is canonical.** 2. **Do not delete anything yet** — v1 (`/v1/*` legacy namespace) stays in place. 3. **Do not refactor old code** unless it supports current or planned functionality. This plan touches **v2 + the shared edge-function backend only.**

## Objective
Make v2 a **maintainable, viable** app that gives users the **best UX solving their problem**, where **technical errors are identified, logged, and fixed fast** — sequenced to serve the **Alpha -> Beta -> GA** roadmap.

## Governing constraint
Refactor spend **<=20-30% of any 20h sprint (<=4-6h)** unless an item is massive-impact (justified inline). **Sprint v3 (now) is refactor-frozen.** Everything schedules into the next sprints' buffer / Always-On.

## The insight driving the abstractions
Nearly every roadmap feature reduces to **four primitives**: (1) conversation -> structured fields; (2) local-first **edit-aware** persistence; (3) **Trevor-voiced LLM artifact**; (4) **feedback moment**. Build those four once, in v2/shared.

## Verified findings (v2 + shared)
| # | Finding | Evidence | Stakes |
|---|---|---|---|
| 1 | Routing confirms v2 is canonical | `src/config/routes.ts` namespaces all legacy under `/v1/*`; `App.tsx` `VersionGate` -> `/v2/coach` is the live entry | v1 is safely isolated — freeze, don't touch |
| 2 | Field-extraction duplicated in v2 | `v2/components/EnhancedChatInterface.tsx`, V2ChatPanel inline regex, `hooks/useFieldExtraction.ts` (474L) | Multiplies with every extracting feature |
| 3 | v2 field persistence over a good spine, but hooks duplicate it | `hooks/{useAvatarFieldSync,useFieldDatabaseSync,useSimpleFieldSync,useFieldSync}` over `lib/knowledge-base/*` (344L+429L) | **Trust foundation** — Alpha gate + QA: "manual edits never overwritten" |
| 4 | Observability partial | shared `components/ErrorBoundary.tsx`; `Result<T>` in v2; **no `reportError`/Sentry**; quota errors hide in HTTP-200 SSE bodies | Blocks Alpha gate; user's #1 goal |
| 5 | Live `/v2/coach` state hook mid-extraction | `hooks/v2/useBrandCoachV2State` + `useChatOrchestration/useFieldOrchestration/useBrandCoachUI` | Current functionality stuck "worst of both" |

## Target abstractions (v2/shared; all ACCEPTED, #7 RESOLVED -> v2)
1. **Single v2 Field-Extraction pipeline** *(RF-05)* 2. **One v2 Field Repository, edit-aware** on the knowledge-base spine *(RF-03)* 3. **Typed LLM edge-fn client** (shared) *(RF-01)* 4. **`Result<T>` + `reportError` sink** *(RF-02)* 5. **Enforce DI in v2** *(RF-07)* 6. **`FeedbackMoment` primitive** *(RF-06)* 7. **v1/v2 canonical -> RESOLVED v2;** in-scope action = finish the v2 coach extraction *(RF-04)*.

## Priority-ordered backlog (full detail in the workbook)
- **P0 RF-01** Typed LLM edge-fn client — *massive-impact exception* (Alpha gate + #1 goal + de-risks 12+ edge fns)
- **P0 RF-02** reportError sink + Result<T> propagation
- **P1 RF-03** Converge v2 Field Repository + single edit-lock (highest correctness risk)
- **P1 RF-04** Finish `useBrandCoachV2State` extraction (live coach — current functionality)
- **P2 RF-05** Unify v2 Field-Extraction · **P2 RF-06** FeedbackMoment primitive
- **P3 RF-07** Enforce DI in v2 · **P4 RF-08** Collapse v2 cosmetic dupes

## Sequencing (<=30% per sprint)
- **Sprint v3 (now):** frozen — 0h.
- **+1 (06-01):** RF-02 (2-3h) + RF-01 thin slice (3-4h).
- **+2 (06-08):** RF-04 finish coach extraction (3-5h) + RF-07 incremental.
- **+3 (06-15):** RF-03 pt1 (4h) + RF-06 (2h).
- **+4 (06-22):** RF-03 pt2 (3h) + RF-05 (3h).
- **Always-On:** RF-07 touches + RF-08, opportunistic.

## V1 disposition — **NO ACTION NOW**
**KEEP (load-bearing):** `/free-diagnostic`, `/beta*`, `/v1/canvas`. **FREEZE (leave untouched):** the rest of `/v1/*` (idea modules, dashboard, journey, avatar, copy-generator, research, conversations) + v1-only services/hooks. **FUTURE (deferred):** decommission is a real LOC win, triggered only once v2 covers the KEEP surfaces and `/v1/*` traffic ~0 — then a measured, test-guarded removal as its own task. Full table in the workbook.

## Observability plan (identify -> log -> fix)
Edge-fn envelope classification (RF-01) -> `Result<T>` at v2 boundaries -> single `reportError` sink (console + dev overlay, **Sentry-ready seam**) -> ErrorBoundary wiring at route level -> Supabase reachability ping/banner. **DoD:** walk the flow with Supabase paused AND a forced quota error; both yield a clear recoverable error + a captured report.

## Out of scope (now)
**Deleting/refactoring any v1 code** (steering), full Sentry/APM (Beta), edge-fn internal rewrites, standalone test backfill (do inline), design-system cleanup, perf work (no measured problem).

---
*Generated by `scripts/build_refactor_plan.py`. Companion workbook: `refactor_build_plan_v1_2026-05-26.xlsx`.*
"""
with open(MD, "w", encoding="utf-8") as f:
    f.write(md)
print("report   ->", MD)
