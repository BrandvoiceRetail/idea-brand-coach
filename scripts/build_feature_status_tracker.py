"""Build docs/feature_status_tracker.xlsx.

Operational counterpart to the launch roadmap workbook. Produces six sheets
(Summary, Features, Sub-features, Phase Gate, Risks, Change Log) with formulas,
data validation, and conditional formatting.

Run: python3 scripts/build_feature_status_tracker.py
"""

from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

OUTPUT_PATH = (
    Path(__file__).resolve().parent.parent / "docs" / "feature_status_tracker.xlsx"
)

# Gen 3 design tokens
COLOR_BLK = "FF111111"
COLOR_GLD = "FFD4960A"
COLOR_GLD_LT = "FFFEF5DC"
COLOR_PILLAR_I = "FFF0B429"
COLOR_PILLAR_D = "FF4CAE53"
COLOR_PILLAR_E = "FF2B9FD4"
COLOR_PILLAR_A = "FFE8850A"

# Status palette
STATUS_FILLS = {
    "complete": PatternFill("solid", fgColor="FFC6EFCE"),
    "in_review": PatternFill("solid", fgColor="FFBDD7EE"),
    "in_progress": PatternFill("solid", fgColor="FFFFEB9C"),
    "blocked": PatternFill("solid", fgColor="FFF8CBAD"),
    "not_started": PatternFill("solid", fgColor="FFD9D9D9"),
}

# Enum vocabularies
STATUS_OPTIONS = ["not_started", "in_progress", "blocked", "in_review", "complete"]
PHASE_OPTIONS = ["Alpha", "Beta", "GA", "Post-GA"]
PRIORITY_OPTIONS = ["P0", "P1", "P2"]
AREA_OPTIONS = [
    "conversation",
    "avatar_builder",
    "canvas",
    "create_mode",
    "coach_mode",
    "pay_gate",
    "export",
    "accounts",
    "asset_tracking",
    "feedback_loop",
    "analytics",
    "infrastructure",
    "design_system",
]
RISK_STATE_OPTIONS = ["monitoring", "triggered", "mitigated", "resolved"]
SEVERITY_OPTIONS = ["Critical", "High", "Medium", "Low"]

HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFFFF")
HEADER_FILL = PatternFill("solid", fgColor=COLOR_BLK)
HEADER_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
BODY_FONT = Font(name="Arial", size=10)
BODY_ALIGN = Alignment(vertical="top", wrap_text=True)
TITLE_FONT = Font(name="Arial", size=16, bold=True, color="FFFFFFFF")
TITLE_FILL = PatternFill("solid", fgColor=COLOR_GLD)
SECTION_FONT = Font(name="Arial", size=11, bold=True, color=COLOR_BLK)
SECTION_FILL = PatternFill("solid", fgColor=COLOR_GLD_LT)

THIN = Side(border_style="thin", color="FFBFBFBF")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

TODAY = date(2026, 5, 13)


# ---------------------------------------------------------------------------
# Seed data
# ---------------------------------------------------------------------------

# Feature inventory. Owner defaults to "Matthew" where uncertain. Effort estimates
# are intentionally rough — the workbook is the place where these get refined.
# (feature_id, name, area, description, phase, priority, status, percent, blocker,
#  owner, depends_on, effort_h, actual_h, notes)
FEATURES: list[tuple] = [
    # --- Alpha P0 ---
    (
        "F-001",
        "Layer 1 conversation engine",
        "conversation",
        "10-question Layer 1 flow with Trevor-voice fallback responses and real-time canvas-sidebar updates.",
        "Alpha",
        "P0",
        "in_progress",
        50,
        "",
        "Matthew",
        "F-018",
        40,
        0,
        "Existing chat infrastructure (SupabaseChatService) is the substrate. Needs Trevor-voice prompt set + 10-Q question script wired.",
    ),
    (
        "F-002",
        "Canvas sidebar real-time population",
        "canvas",
        "Canvas sidebar updates as Layer 1 conversation progresses, showing extracted fields with confidence indicators.",
        "Alpha",
        "P0",
        "in_progress",
        40,
        "",
        "Matthew",
        "F-001,F-012",
        20,
        0,
        "Field extraction pipeline exists; sidebar render needs to subscribe to extraction events.",
    ),
    (
        "F-003",
        "Draft Canvas synthesis",
        "canvas",
        "Synthesise Layer 1 inputs into a coherent Draft Canvas with DRAFT badge and clear Forensic upgrade prompt.",
        "Alpha",
        "P0",
        "in_progress",
        45,
        "",
        "Matthew",
        "F-001,F-018",
        24,
        0,
        "Edge function exists for canvas synthesis. DRAFT badge + Forensic upgrade CTA needs design pass.",
    ),
    (
        "F-004",
        "Avatar Builder — Psychographic Profile",
        "avatar_builder",
        "Section 1 of the four-section Avatar Builder. Captures values, identity, worldview.",
        "Alpha",
        "P0",
        "not_started",
        0,
        "",
        "Matthew",
        "F-003,F-009",
        32,
        0,
        "Gen 3 spec is the source of truth. Existing V2 PRD predates Gen 3; verify field list.",
    ),
    (
        "F-005",
        "Avatar Builder — Buying Behaviour",
        "avatar_builder",
        "Section 2. Captures purchase triggers, objections, decision criteria, channels.",
        "Alpha",
        "P0",
        "not_started",
        0,
        "",
        "Matthew",
        "F-004",
        32,
        0,
        "",
    ),
    (
        "F-006",
        "Avatar Builder — Emotional Journey Mapping",
        "avatar_builder",
        "Section 3. Pre/during/post purchase emotional states; the Trigger Moment originates here.",
        "Alpha",
        "P0",
        "not_started",
        0,
        "",
        "Matthew",
        "F-005",
        36,
        0,
        "Trigger Moment is Trevor's flagged most-important field — must surface prominently.",
    ),
    (
        "F-007",
        "Avatar Builder — Voice of Customer",
        "avatar_builder",
        "Section 4. Paste-in reviews and testimonials; extract verbatim language for copy hooks.",
        "Alpha",
        "P0",
        "not_started",
        0,
        "",
        "Matthew",
        "F-006",
        40,
        0,
        "Requires voice-of-customer-analysis edge function.",
    ),
    (
        "F-008",
        "Forensic Canvas synthesis",
        "canvas",
        "28+ field canvas synthesised from Avatar Builder. Trigger Moment displayed as 'Strategic Entry Point'.",
        "Alpha",
        "P0",
        "not_started",
        0,
        "",
        "Matthew",
        "F-004,F-005,F-006,F-007",
        40,
        0,
        "Differential vs Draft Canvas is the core demonstration of value.",
    ),
    (
        "F-009",
        "AI Assistance system",
        "avatar_builder",
        "Four modes — Think About This, Real-World Examples, Get AI Suggestions, Get AI Auto-Populate — available per field.",
        "Alpha",
        "P0",
        "in_progress",
        30,
        "",
        "Matthew",
        "F-018",
        48,
        0,
        "AIAssistant component exists for V2 fields; needs Gen 3 mode expansion.",
    ),
    (
        "F-010",
        "Progressive disclosure UI",
        "design_system",
        "Avatar Builder fields reveal progressively to avoid form fatigue; mobile-first.",
        "Alpha",
        "P0",
        "in_progress",
        25,
        "",
        "Matthew",
        "",
        24,
        0,
        "Existing V2 UI is form-shaped. Trevor's Risk Register flag: 'feels like a form, not a conversation' is Critical.",
    ),
    (
        "F-011",
        "Manual edit lock",
        "avatar_builder",
        "Fields manually edited by the user are flagged and never silently overwritten by AI suggestions.",
        "Alpha",
        "P0",
        "in_progress",
        60,
        "",
        "Matthew",
        "",
        12,
        0,
        "Trust foundation per Trevor + existing PRD constraint. Partially present in V2 quality badge code.",
    ),
    (
        "F-012",
        "Field extraction pipeline with confidence scoring",
        "infrastructure",
        "Conversation inputs → structured field extraction with per-field confidence score. Drives the quality badges.",
        "Alpha",
        "P0",
        "in_progress",
        55,
        "",
        "Matthew",
        "",
        32,
        0,
        "Field extraction shipped recently (commit 77ae116 — IDEA Brand Coach quality score badges).",
    ),
    (
        "F-013",
        "Feedback Moment 1 — Entry signal modal",
        "feedback_loop",
        "Modal after Draft Canvas asks whether the conversation revealed new territory; routes to feedback sheet.",
        "Alpha",
        "P0",
        "not_started",
        0,
        "",
        "Matthew",
        "F-003,F-017",
        12,
        0,
        "Spec lives in Launch Roadmap › Feedback Loop sheet.",
    ),
    (
        "F-014",
        "Feedback Moment 2 — Use intent inline panel",
        "feedback_loop",
        "Inline panel after Forensic Canvas asks what the user intends to do with it (and surfaces asset-update demand signal).",
        "Alpha",
        "P0",
        "not_started",
        0,
        "",
        "Matthew",
        "F-008,F-017",
        12,
        0,
        "% selecting 'update existing assets' is the Alpha→Beta asset-tracking gate.",
    ),
    (
        "F-015",
        "Feedback Moment 3 — 14-day async email",
        "feedback_loop",
        "Day-14 email asks what the user has actually done with the canvas and what's missing.",
        "Alpha",
        "P0",
        "not_started",
        0,
        "",
        "Matthew",
        "F-017",
        8,
        0,
        "Requires scheduled-send infrastructure (Resend / Supabase cron).",
    ),
    (
        "F-016",
        "Feedback response sheet integration",
        "infrastructure",
        "All three feedback moments write to a single sheet (Google Sheets or Supabase view) Matthew can monitor daily.",
        "Alpha",
        "P0",
        "not_started",
        0,
        "",
        "Matthew",
        "",
        8,
        0,
        "Single destination simplifies the alpha findings synthesis step.",
    ),
    (
        "F-017",
        "Error monitoring",
        "infrastructure",
        "Sentry (or equivalent) capturing client + edge function errors. Daily review during alpha.",
        "Alpha",
        "P0",
        "not_started",
        0,
        "",
        "Matthew",
        "",
        6,
        0,
        "Cheap to install. Risk Register flags that testers won't report half the bugs.",
    ),
    (
        "F-018",
        "Tester onboarding flow",
        "accounts",
        "One-page Loom + in-app welcome that frames alpha state and the three feedback moments.",
        "Alpha",
        "P0",
        "not_started",
        0,
        "",
        "Matthew",
        "",
        8,
        0,
        "No tutorial. Sets expectation that this is alpha and feedback is the deliverable.",
    ),
    # --- Beta P0 ---
    (
        "F-019",
        "Pay gate",
        "pay_gate",
        "Stripe integration with two pricing options (£67 one-time, £29/mo) and post-payment action menu.",
        "Beta",
        "P0",
        "not_started",
        0,
        "",
        "Matthew",
        "F-008",
        40,
        0,
        "Central conversion test of beta. Pay-gate previews must use real canvas data.",
    ),
    (
        "F-020",
        "Create mode — Amazon listing copy",
        "create_mode",
        "Generate Amazon listing copy from Forensic Canvas using Trevor-voice prompts.",
        "Beta",
        "P0",
        "not_started",
        0,
        "",
        "Matthew",
        "F-008,F-019",
        32,
        0,
        "First Create-mode output. Anchors the pay-gate preview.",
    ),
    (
        "F-021",
        "Create mode — Brand Voice Guide",
        "create_mode",
        "Generate Brand Voice Guide document from Forensic Canvas.",
        "Beta",
        "P0",
        "not_started",
        0,
        "",
        "Matthew",
        "F-008,F-019",
        24,
        0,
        "",
    ),
    (
        "F-022",
        "Coach mode v1",
        "coach_mode",
        "Open conversation with full canvas context and document upload support.",
        "Beta",
        "P0",
        "not_started",
        0,
        "",
        "Matthew",
        "F-008",
        56,
        0,
        "Required to deliver the £29/mo subscription value prop per Trevor's spec.",
    ),
    (
        "F-023",
        "Export — PDF and Word",
        "export",
        "PDF and Word exports of both Draft and Forensic canvases.",
        "Beta",
        "P0",
        "not_started",
        0,
        "",
        "Matthew",
        "F-003,F-008",
        24,
        0,
        "Required to deliver the canvas-as-artifact promise.",
    ),
    (
        "F-024",
        "Analytics + funnel tracking",
        "analytics",
        "PostHog or Mixpanel events on every funnel step; cohort analysis enabled.",
        "Beta",
        "P0",
        "not_started",
        0,
        "",
        "Matthew",
        "",
        16,
        0,
        "Conversion data is the whole point of beta.",
    ),
    (
        "F-025",
        "Support workflow",
        "infrastructure",
        "Intercom (or email) inbox with documented SLA; paying users get answers.",
        "Beta",
        "P0",
        "not_started",
        0,
        "",
        "Matthew",
        "",
        8,
        0,
        "Refund-rate risk is partly a support-responsiveness problem.",
    ),
    (
        "F-026",
        "ROI testimonial harvest email sequence",
        "feedback_loop",
        "Day-0, day-14, day-30, day-60 emails plus debrief invite when results report positive.",
        "Beta",
        "P0",
        "not_started",
        0,
        "",
        "Matthew",
        "F-015",
        12,
        0,
        "Target: 5+ verified testimonials by end of Beta; 10+ within 3 months.",
    ),
    (
        "F-027",
        "Brand Diagnostic integration",
        "conversation",
        "Diagnostic CTA dynamic (Brand Coach vs strategist) + score handoff into Layer 1 + visual identity refresh.",
        "Beta",
        "P0",
        "in_progress",
        70,
        "",
        "Matthew",
        "",
        20,
        0,
        "Diagnostic largely exists (FreeDiagnostic + BetaFeedback components). Handoff into Layer 1 still loose.",
    ),
    # --- Beta conditional (asset tracking) ---
    (
        "F-028",
        "Asset Inventory v1",
        "asset_tracking",
        "User can add assets, tag type (listing/email/social/ad/landing/product/other), mark alignment status manually.",
        "Beta",
        "P1",
        "not_started",
        0,
        "Conditional on Alpha decision gate",
        "Matthew",
        "F-014",
        40,
        0,
        "DEFERRED DECISION. Gate at end of Alpha. Default is NO. Trevor's Gen 3 spec does not include asset tracking — only ship if Moment 2 signal is ≥40%.",
    ),
    (
        "F-029",
        "Asset Inventory v1.5 — Canvas linkage",
        "asset_tracking",
        "Mark asset as aligned to a specific canvas version; later canvas updates flag asset as stale.",
        "Beta",
        "P1",
        "not_started",
        0,
        "Conditional on F-028",
        "Matthew",
        "F-028",
        32,
        0,
        "Introduces the temporal signal — user can see what's out of date.",
    ),
    (
        "F-030",
        "Asset Inventory v2 — Diff surface",
        "asset_tracking",
        "When canvas changes, show user a list of likely-affected assets with diff reason (Trigger Moment changed, voice changed, etc.).",
        "GA",
        "P1",
        "not_started",
        0,
        "Conditional on F-029",
        "Matthew",
        "F-029",
        56,
        0,
        "Likely a GA-era feature per brief; listed for completeness.",
    ),
    # --- GA P0 ---
    (
        "F-031",
        "My Brands dashboard",
        "accounts",
        "Multi-brand dashboard with brand switcher; users can manage 2+ brands.",
        "GA",
        "P0",
        "not_started",
        0,
        "",
        "Matthew",
        "F-046",
        40,
        0,
        "Power users will have multiple brands; product caps out at one customer without this.",
    ),
    (
        "F-032",
        "Anonymous-to-authenticated migration",
        "accounts",
        "Diagnostic completers without accounts have their data preserved on signup.",
        "GA",
        "P0",
        "not_started",
        0,
        "",
        "Matthew",
        "F-027",
        16,
        0,
        "Top-of-funnel users drop off without this.",
    ),
    (
        "F-033",
        "Self-serve onboarding tour",
        "design_system",
        "Tooltip-light onboarding tour; no human needed to start.",
        "GA",
        "P0",
        "not_started",
        0,
        "",
        "Matthew",
        "",
        16,
        0,
        "Public launch means support cost scales with users — self-serve protects margin.",
    ),
    (
        "F-034",
        "Pricing page (cold-traffic optimised)",
        "pay_gate",
        "Standalone pricing page that converts cold traffic; A/B tested in late beta.",
        "GA",
        "P0",
        "not_started",
        0,
        "",
        "Matthew",
        "F-019",
        16,
        0,
        "Conversion surface for non-warm visitors.",
    ),
    (
        "F-035",
        "Status page",
        "infrastructure",
        "Public uptime page + incident communication channel.",
        "GA",
        "P0",
        "not_started",
        0,
        "",
        "Matthew",
        "",
        8,
        0,
        "Trust signal required for paid users.",
    ),
    (
        "F-036",
        "Knowledge base",
        "infrastructure",
        "10–20 articles seeded from Beta support tickets covering common questions.",
        "GA",
        "P0",
        "not_started",
        0,
        "",
        "Matthew",
        "F-025",
        24,
        0,
        "Reduces ticket volume and improves time-to-value.",
    ),
    (
        "F-037",
        "Refund policy + ToS + Privacy Policy",
        "infrastructure",
        "Standard SaaS refund (14-day), Terms of Service, Privacy Policy.",
        "GA",
        "P0",
        "not_started",
        0,
        "",
        "Matthew",
        "",
        8,
        0,
        "Legal hygiene required for paid public product.",
    ),
    (
        "F-038",
        "Performance metrics tracking",
        "analytics",
        "Per-avatar manual entry of ROAS, CTR, CPA, conversion rate, custom metrics. Drives ROI testimonials.",
        "GA",
        "P0",
        "not_started",
        0,
        "",
        "Matthew",
        "F-046",
        32,
        0,
        "OPEN QUESTION: existing V2 PRD scopes this as a P0 V2 feature. Trevor's Gen 3 spec does not address it. Treated here as GA P0; needs Matthew sign-off.",
    ),
    # --- GA P1 ---
    (
        "F-039",
        "Referral programme",
        "infrastructure",
        "In-product referral mechanism with 20% rev share.",
        "GA",
        "P1",
        "not_started",
        0,
        "",
        "Matthew",
        "F-019",
        24,
        0,
        "Acquisition lever from GA Plan.",
    ),
    (
        "F-040",
        "Affiliate tracking infrastructure",
        "infrastructure",
        "Partner affiliate links, attribution, payout reporting.",
        "GA",
        "P1",
        "not_started",
        0,
        "",
        "Matthew",
        "",
        24,
        0,
        "Brand strategist partner channel needs attribution before outreach.",
    ),
    (
        "F-041",
        "SEO foundation",
        "infrastructure",
        "Sitemap, schema markup, content templates for long-tail brand strategy queries.",
        "GA",
        "P1",
        "not_started",
        0,
        "",
        "Matthew",
        "",
        20,
        0,
        "Three-month organic ramp — start at beginning of beta to earn GA traffic.",
    ),
    (
        "F-042",
        "Avatar comparison view",
        "avatar_builder",
        "Side-by-side comparison of two avatars (different segments, different brands).",
        "GA",
        "P1",
        "not_started",
        0,
        "",
        "Matthew",
        "F-031",
        24,
        0,
        "Power-user feature; rewards multi-canvas usage.",
    ),
    (
        "F-043",
        "Field edit history viewer",
        "avatar_builder",
        "Audit log per field showing edit history and which canvas version each value was active in.",
        "GA",
        "P1",
        "not_started",
        0,
        "",
        "Matthew",
        "F-046",
        16,
        0,
        "Schema needs field_edit_history table.",
    ),
    (
        "F-044",
        "Bulk operations",
        "avatar_builder",
        "Copy fields between avatars; find/replace across an avatar.",
        "GA",
        "P1",
        "not_started",
        0,
        "",
        "Matthew",
        "F-031",
        20,
        0,
        "Power-user feature.",
    ),
    # --- GA + TikTok ---
    (
        "F-045a",
        "TikTok-specific landing page",
        "design_system",
        "Variant landing page optimised for TikTok cold traffic.",
        "Post-GA",
        "P1",
        "not_started",
        0,
        "",
        "Matthew",
        "F-034",
        12,
        0,
        "TikTok integration is GA+4wks per Launch Roadmap; not a build dependency.",
    ),
    (
        "F-045b",
        "UTM tagging across TikTok touchpoints",
        "analytics",
        "Consistent UTM scheme on every TikTok-originating link.",
        "Post-GA",
        "P1",
        "not_started",
        0,
        "",
        "Matthew",
        "F-024",
        4,
        0,
        "",
    ),
    (
        "F-045c",
        "TikTok Pixel integration",
        "analytics",
        "TikTok Pixel installed on landing pages and conversion events.",
        "Post-GA",
        "P1",
        "not_started",
        0,
        "",
        "Matthew",
        "",
        4,
        0,
        "",
    ),
    (
        "F-045d",
        "Self-report attribution question on signup",
        "analytics",
        "'Where did you hear about us?' on signup; option for TikTok and named creators.",
        "Post-GA",
        "P1",
        "not_started",
        0,
        "",
        "Matthew",
        "",
        4,
        0,
        "Pairs with UTM data to triangulate true attribution.",
    ),
    # --- Infrastructure (cross-phase) ---
    (
        "F-046",
        "Database migrations (additive)",
        "infrastructure",
        "Migrations for brands, avatars, avatar_profiles, brand_strategies, performance_metrics, field_edit_history, assets (if asset tracking ships).",
        "Alpha",
        "P0",
        "in_progress",
        50,
        "",
        "Matthew",
        "",
        24,
        0,
        "Additive only. RLS policies in F-047.",
    ),
    (
        "F-047",
        "RLS policies for new tables",
        "infrastructure",
        "Row-level-security policies on every new table created in F-046.",
        "Alpha",
        "P0",
        "in_progress",
        40,
        "",
        "Matthew",
        "F-046",
        12,
        0,
        "",
    ),
    (
        "F-048",
        "Edge functions",
        "infrastructure",
        "conversation, canvas-synthesis, create-copy, coach-conversation, voice-of-customer-analysis edge functions.",
        "Alpha",
        "P0",
        "in_progress",
        40,
        "",
        "Matthew",
        "",
        40,
        0,
        "idea-framework-consultant exists; Gen 3 functions partially overlap. Audit needed.",
    ),
    (
        "F-049",
        "RAG integration with System KB",
        "infrastructure",
        "Retrieval from the existing System KB vector store (vs_6948707b318c81918a90e9b44970a99e) in all relevant edge functions.",
        "Alpha",
        "P0",
        "in_progress",
        60,
        "",
        "Matthew",
        "F-048",
        16,
        0,
        "OpenAI embeddings removed in commit 260128b (consolidated to Claude/internal embeddings) — verify RAG pipeline still wired.",
    ),
    (
        "F-050",
        "Feature flag system",
        "infrastructure",
        "Ensure useFeatureFlag gates every V2/Gen 3 route; flags toggle independently.",
        "Alpha",
        "P0",
        "in_progress",
        80,
        "",
        "Matthew",
        "",
        8,
        0,
        "Hook already exists. Audit V2 routes for missing gates.",
    ),
    (
        "F-051",
        "Pre-implementation refactoring sprint",
        "infrastructure",
        "Week-0 refactoring sprint per V2 PRD: service extraction, field sync architecture, component composition.",
        "Alpha",
        "P0",
        "in_progress",
        70,
        "",
        "Matthew",
        "",
        40,
        0,
        "Refactor patterns shipped (SupabaseChatService split, FieldSyncService, AIAssistant decomposition). Verify all Gen 3 work picks up new patterns.",
    ),
    # --- Power-user / multi-variant features (added 2026-05-16 per Matthew) ---
    (
        "F-053",
        "Multiple positioning angles per avatar",
        "avatar_builder",
        "Support multiple positioning / marketing angle variants for the same avatar — each angle has its own brand strategy doc / canvas section, but shares the underlying avatar.",
        "Beta",
        "P1",
        "not_started",
        0,
        "",
        "Matthew",
        "F-008,F-031",
        40,
        0,
        "Validates A/B angle testing for paying users. Each angle = a strategy variant tied to a canonical avatar. Surfaces in Coach mode and Export. OPEN: schema choice — separate canvases per angle vs angle-tagged sections within one canvas.",
    ),
    (
        "F-054",
        "Image asset side-by-side comparison and tracking",
        "asset_tracking",
        "Upload current image/video assets, view side-by-side, tag against avatar + angle, flag gaps and emotional-connection opportunities.",
        "GA",
        "P1",
        "not_started",
        0,
        "Conditional on F-028 (asset tracking decision)",
        "Matthew",
        "F-028,F-053",
        56,
        0,
        "Extension of asset_tracking that goes beyond text. Anchored to InfinityVault need N-002 (assess images/videos, identify what's missing, surface emotional-connection opportunities). Requires image storage (Supabase Storage) and a comparison UI.",
    ),
    (
        "F-055",
        "Design brief creation",
        "create_mode",
        "Generate a design brief document from the Forensic Canvas + selected positioning angle. Output briefs creative work (image, video, packaging) for designers or agencies.",
        "Beta",
        "P1",
        "not_started",
        0,
        "",
        "Matthew",
        "F-008,F-053",
        32,
        0,
        "Sibling to F-020 (Amazon listing copy) and F-021 (Brand Voice Guide). Bridges the canvas to creative work. Required to make F-054 actionable (gap → brief → asset).",
    ),
    (
        "F-056",
        "Image generation (coming soon)",
        "create_mode",
        "Generate brand-aligned imagery directly from canvas + design brief. Marked 'coming soon' — Post-GA feature, not in active build.",
        "Post-GA",
        "P2",
        "not_started",
        0,
        "",
        "Matthew",
        "F-055",
        80,
        0,
        "Surfaced as 'coming soon' in UI (teaser). Closes the loop from canvas → design brief → generated imagery. No model choice locked in; defer until Post-GA stability work is done.",
    ),
    (
        "F-052",
        "Lovable.dev vs existing codebase decision",
        "infrastructure",
        "Architectural decision: build Gen 3 in Lovable.dev (per Trevor) or extend existing React codebase. Critical risk per Launch Roadmap.",
        "Alpha",
        "P0",
        "blocked",
        0,
        "Awaiting decision — flagged Critical in Risk Register",
        "Matthew",
        "",
        4,
        0,
        "DEFERRED DECISION. Trevor's Gen 3 Brief specifies Lovable; existing codebase is React + Supabase. Don't half-port. Default: extend existing repo until Matthew decides otherwise.",
    ),
]

# Sub-features for the Alpha-required P0 features. Each entry is:
# (subfeature_id, parent_feature_id, name, description, acceptance_criteria,
#  status, percent, effort_h, actual_h, blocker, notes)
SUBFEATURES: list[tuple] = [
    # F-001 Layer 1 conversation engine
    (
        "F-001.1",
        "F-001",
        "10-question script in Trevor's voice",
        "Source the 10 Layer 1 questions and Trevor-voice phrasing from IDEA_KNOWLEDGE.md.",
        "All 10 questions stored in prompt config and rendered in order; Trevor reviews sample transcript.",
        "in_progress",
        60,
        8,
        0,
        "",
        "Pull canonical wording from Trevor's Knowledge doc.",
    ),
    (
        "F-001.2",
        "F-001",
        "Fallback responses for empty / off-topic input",
        "When the user's reply doesn't extract a field, surface a Trevor-voice nudge instead of restating the question.",
        "Off-topic input triggers fallback string; fallbacks vary by question; no robotic 'Please answer the question.'",
        "in_progress",
        40,
        6,
        0,
        "",
        "",
    ),
    (
        "F-001.3",
        "F-001",
        "Conversation state persistence",
        "Layer 1 progress persists across reload via existing FieldSyncService.",
        "Refreshing mid-conversation resumes at the last unanswered question with prior answers visible.",
        "in_progress",
        70,
        6,
        0,
        "",
        "Leverages existing local-first sync.",
    ),
    (
        "F-001.4",
        "F-001",
        "Streaming response rendering",
        "Assistant replies stream token-by-token via SSE.",
        "User sees streaming text within 500ms of submit; SSE error states render a Trevor-voice fallback.",
        "in_progress",
        50,
        8,
        0,
        "",
        "Capture raw SSE bytes to distinguish frontend fallback from edge function errors (see feedback_sse_stream_capture).",
    ),
    (
        "F-001.5",
        "F-001",
        "Mobile keyboard / viewport handling",
        "Conversation UI works on mobile without keyboard-induced layout breaks.",
        "iOS Safari and Android Chrome both render correctly with keyboard open; input stays visible.",
        "not_started",
        0,
        6,
        0,
        "",
        "Mobile-primary testers are >=50% of the alpha pool.",
    ),
    # F-002 Canvas sidebar
    (
        "F-002.1",
        "F-002",
        "Subscribe sidebar to field extraction events",
        "Sidebar listens for new/updated field extractions and rerenders affected cards.",
        "Extracting a field during conversation updates the corresponding sidebar card within 1s.",
        "in_progress",
        50,
        8,
        0,
        "",
        "",
    ),
    (
        "F-002.2",
        "F-002",
        "Confidence-score badges on sidebar fields",
        "Each populated sidebar field shows the quality / confidence badge (already shipped for V2 — port to Layer 1).",
        "Low-confidence fields show amber badge; high-confidence shows green; matches existing badge design.",
        "complete",
        100,
        4,
        4,
        "",
        "Quality badges shipped in commit 77ae116.",
    ),
    (
        "F-002.3",
        "F-002",
        "Sidebar collapse / expand on mobile",
        "Mobile users can collapse sidebar to focus on conversation.",
        "Hamburger toggle hides sidebar; state persists across navigation.",
        "not_started",
        0,
        4,
        0,
        "",
        "",
    ),
    # F-003 Draft Canvas synthesis
    (
        "F-003.1",
        "F-003",
        "DRAFT badge component",
        "Visual badge on Draft Canvas distinguishing it from Forensic Canvas.",
        "Badge visible above-fold on canvas page; uses gold / amber tone from design tokens.",
        "not_started",
        0,
        4,
        0,
        "",
        "",
    ),
    (
        "F-003.2",
        "F-003",
        "Synthesis edge function with Trevor prompt",
        "canvas-synthesis edge function takes Layer 1 transcript + extracted fields, returns Draft Canvas JSON.",
        "Returns all Draft-tier fields populated; fields with low confidence flagged; latency <8s.",
        "in_progress",
        60,
        12,
        0,
        "",
        "Existing function present; needs Gen 3 prompt update.",
    ),
    (
        "F-003.3",
        "F-003",
        "Forensic upgrade CTA",
        "Draft Canvas surfaces a prominent CTA inviting user into the Avatar Builder.",
        "CTA visible in canvas top-bar AND inline after the most synthesis-light section.",
        "not_started",
        0,
        4,
        0,
        "",
        "",
    ),
    (
        "F-003.4",
        "F-003",
        "Coherence pass",
        "Synthesis avoids contradictions between fields (e.g., aspirational language in one field vs jaded in another).",
        "Manual review of 5 sample draft canvases shows no contradictions; Trevor signs off.",
        "not_started",
        0,
        4,
        0,
        "",
        "",
    ),
    # F-004 Avatar Builder — Psychographic Profile
    (
        "F-004.1",
        "F-004",
        "Field schema for Psychographic Profile",
        "Define the field set from Gen 3 spec (values, identity, worldview, lifestyle).",
        "Schema documented and reviewed against Gen 3 brief; types defined in TypeScript.",
        "not_started",
        0,
        4,
        0,
        "",
        "",
    ),
    (
        "F-004.2",
        "F-004",
        "Section UI with progressive disclosure",
        "Section renders one cluster at a time, expanding as the user advances.",
        "User sees no more than ~4 fields at once on mobile.",
        "not_started",
        0,
        12,
        0,
        "",
        "",
    ),
    (
        "F-004.3",
        "F-004",
        "AI assistance modes wired",
        "Each field supports the four AI assistance modes from F-009.",
        "All four mode buttons present per field; each routes to the right edge function.",
        "not_started",
        0,
        8,
        0,
        "",
        "Depends on F-009.",
    ),
    (
        "F-004.4",
        "F-004",
        "Persistence and resume",
        "Section state persists via FieldSyncService and resumes on reload.",
        "Closing browser mid-section and reopening restores all entered values.",
        "not_started",
        0,
        4,
        0,
        "",
        "",
    ),
    (
        "F-004.5",
        "F-004",
        "Completion gate to next section",
        "User can advance to Section 2 once Section 1 is in a valid state (not necessarily 100% — Trevor wants flow, not gates).",
        "Advance button enables when required-minimum fields are filled; soft prompt for empty optional fields.",
        "not_started",
        0,
        4,
        0,
        "",
        "Gen 3 spec discourages hard gates; mirror that here.",
    ),
    # F-005
    (
        "F-005.1",
        "F-005",
        "Field schema for Buying Behaviour",
        "Define fields: purchase triggers, objections, decision criteria, channels, price sensitivity.",
        "Schema documented; matches Gen 3 brief.",
        "not_started",
        0,
        4,
        0,
        "",
        "",
    ),
    (
        "F-005.2",
        "F-005",
        "Section UI",
        "Progressive disclosure UI for the section, consistent with F-004.",
        "Visual parity with Section 1; mobile-first layout.",
        "not_started",
        0,
        12,
        0,
        "",
        "",
    ),
    (
        "F-005.3",
        "F-005",
        "AI assistance wired",
        "All four AI assistance modes available per field.",
        "All buttons present and functional.",
        "not_started",
        0,
        8,
        0,
        "",
        "",
    ),
    (
        "F-005.4",
        "F-005",
        "Persistence and resume",
        "Section state persists and resumes via FieldSyncService.",
        "Closing and reopening restores state.",
        "not_started",
        0,
        4,
        0,
        "",
        "",
    ),
    # F-006
    (
        "F-006.1",
        "F-006",
        "Field schema for Emotional Journey Mapping",
        "Define fields: pre-purchase emotions, during-purchase emotions, post-purchase emotions, Trigger Moment.",
        "Schema documented; Trigger Moment flagged as the most important.",
        "not_started",
        0,
        4,
        0,
        "",
        "Trevor flagged Trigger Moment as the single most important field.",
    ),
    (
        "F-006.2",
        "F-006",
        "Trigger Moment field surfacing",
        "Trigger Moment field uses elevated visual treatment (gold accent, larger label).",
        "Field visually dominant in section; matches design tokens.",
        "not_started",
        0,
        4,
        0,
        "",
        "",
    ),
    (
        "F-006.3",
        "F-006",
        "Section UI",
        "Progressive disclosure UI consistent with F-004.",
        "Visual parity with prior sections.",
        "not_started",
        0,
        12,
        0,
        "",
        "",
    ),
    (
        "F-006.4",
        "F-006",
        "AI assistance wired",
        "All four AI assistance modes available per field, including specific Trigger-Moment-shaped prompts.",
        "AI buttons functional; Trigger Moment prompts pull example trigger language.",
        "not_started",
        0,
        12,
        0,
        "",
        "",
    ),
    (
        "F-006.5",
        "F-006",
        "Persistence and resume",
        "Section state persists and resumes.",
        "Closing and reopening restores state.",
        "not_started",
        0,
        4,
        0,
        "",
        "",
    ),
    # F-007
    (
        "F-007.1",
        "F-007",
        "Review paste-in UI",
        "User can paste in 5-50 customer reviews / testimonials; UI handles long-form text well.",
        "Textarea accepts pasted content; preview shows reviews chunked and counted.",
        "not_started",
        0,
        8,
        0,
        "",
        "",
    ),
    (
        "F-007.2",
        "F-007",
        "voice-of-customer-analysis edge function",
        "Edge function extracts verbatim copy hooks, common phrases, and emotional language from pasted reviews.",
        "Returns ranked list of phrases with frequency + emotional valence; latency <15s for 25 reviews.",
        "not_started",
        0,
        16,
        0,
        "",
        "Critical for differentiation per Trevor.",
    ),
    (
        "F-007.3",
        "F-007",
        "Extracted phrases UI",
        "Show extracted phrases as cards user can flag as 'use this' or 'discard'.",
        "Flagging persists; flagged phrases populate Forensic Canvas Voice fields.",
        "not_started",
        0,
        8,
        0,
        "",
        "",
    ),
    (
        "F-007.4",
        "F-007",
        "Persistence and resume",
        "Pasted reviews and flagged phrases persist across sessions.",
        "Closing and reopening restores state.",
        "not_started",
        0,
        4,
        0,
        "",
        "",
    ),
    # F-008 Forensic Canvas
    (
        "F-008.1",
        "F-008",
        "28+ field canvas schema",
        "Define every Forensic Canvas field per Gen 3 brief.",
        "Schema documented; field list matches brief.",
        "not_started",
        0,
        4,
        0,
        "",
        "",
    ),
    (
        "F-008.2",
        "F-008",
        "Synthesis edge function (Forensic tier)",
        "canvas-synthesis edge function in Forensic mode pulls from all 4 Avatar Builder sections.",
        "Returns populated Forensic Canvas JSON; differentiated from Draft synthesis.",
        "not_started",
        0,
        16,
        0,
        "",
        "",
    ),
    (
        "F-008.3",
        "F-008",
        "Trigger Moment 'Strategic Entry Point' callout",
        "Forensic Canvas displays Trigger Moment in a prominent callout block labelled Strategic Entry Point.",
        "Callout visible above-fold on canvas page.",
        "not_started",
        0,
        4,
        0,
        "",
        "Trevor flagged this explicitly.",
    ),
    (
        "F-008.4",
        "F-008",
        "Draft-to-Forensic differential view",
        "User can toggle between Draft and Forensic side-by-side to see the upgrade differential.",
        "Toggle visible from canvas; renders both canvases readably on desktop and mobile.",
        "not_started",
        0,
        12,
        0,
        "",
        "Core demonstration of value per Alpha Plan.",
    ),
    (
        "F-008.5",
        "F-008",
        "Export prep (Markdown intermediate)",
        "Forensic Canvas serialises to Markdown as an intermediate format for F-023 PDF/Word export.",
        "Markdown round-trips: re-importing reproduces canvas state.",
        "not_started",
        0,
        4,
        0,
        "",
        "Reuses existing canvas-markdown design (see CANVAS_MARKDOWN_EXPORT_DESIGN.md).",
    ),
    # F-009 AI Assistance
    (
        "F-009.1",
        "F-009",
        "Mode: Think About This",
        "Generates prompt-style question helping the user think more deeply about the field.",
        "Returns 3-5 reflective questions in Trevor's voice; <5s latency.",
        "in_progress",
        40,
        8,
        0,
        "",
        "AIAssistant component scaffolding exists.",
    ),
    (
        "F-009.2",
        "F-009",
        "Mode: Real-World Examples",
        "Returns 2-3 anonymised example fills for the field from RAG over the System KB.",
        "Examples are relevant to the brand category; <8s latency.",
        "in_progress",
        30,
        8,
        0,
        "",
        "",
    ),
    (
        "F-009.3",
        "F-009",
        "Mode: Get AI Suggestions",
        "Returns 2-3 suggested fills based on the user's prior fields.",
        "Suggestions are coherent with prior fields; <8s latency.",
        "in_progress",
        30,
        8,
        0,
        "",
        "",
    ),
    (
        "F-009.4",
        "F-009",
        "Mode: Get AI Auto-Populate",
        "Auto-fills the field with the best inferred value; user can accept / edit / reject.",
        "Auto-fill respects manual edit lock (F-011); never overwrites manual edits silently.",
        "in_progress",
        20,
        12,
        0,
        "",
        "",
    ),
    (
        "F-009.5",
        "F-009",
        "Field-level mode selector UI",
        "Each field shows the 4 mode buttons in a consistent place; tooltip explains each.",
        "Buttons visible and accessible; consistent across all Avatar Builder sections.",
        "in_progress",
        50,
        12,
        0,
        "",
        "AISuggestionHandler component exists.",
    ),
    # F-010 Progressive disclosure
    (
        "F-010.1",
        "F-010",
        "Disclosure pattern documented in design system",
        "Pattern guide written so all sections behave the same way.",
        "Doc in design system; example component referenced.",
        "in_progress",
        20,
        4,
        0,
        "",
        "",
    ),
    (
        "F-010.2",
        "F-010",
        "Mobile-first audit on V2 sections",
        "Audit existing V2 sections against mobile-primary use; document gaps.",
        "Audit report committed; gaps tracked as sub-tasks of relevant features.",
        "in_progress",
        30,
        4,
        0,
        "",
        "",
    ),
    # F-011 Manual edit lock
    (
        "F-011.1",
        "F-011",
        "Manual-edit flag on field model",
        "Persist a boolean per field indicating user has manually edited it.",
        "Flag set on any non-AI edit; survives reload.",
        "complete",
        100,
        4,
        4,
        "",
        "Partially shipped with quality badge work.",
    ),
    (
        "F-011.2",
        "F-011",
        "AI overwrite guard",
        "AI assistance modes never silently overwrite manually edited fields.",
        "Auto-populate skips manual-edited fields by default; user can explicitly opt in to overwrite.",
        "in_progress",
        50,
        4,
        0,
        "",
        "",
    ),
    (
        "F-011.3",
        "F-011",
        "Visual indicator of manual edit",
        "Manually edited fields show a small icon indicating they're locked.",
        "Icon visible; tooltip explains the lock.",
        "in_progress",
        30,
        4,
        0,
        "",
        "",
    ),
    # F-012 Field extraction pipeline
    (
        "F-012.1",
        "F-012",
        "Conversation→field extraction edge function",
        "Edge function takes a conversation turn and returns structured field updates with confidence.",
        "Returns field updates with confidence scores 0-1; latency <3s.",
        "in_progress",
        70,
        12,
        0,
        "",
        "Recent commits indicate substantial work shipped.",
    ),
    (
        "F-012.2",
        "F-012",
        "Confidence score → quality badge mapping",
        "Confidence scores map to the user-facing quality badge tiers.",
        "Mapping documented; consistent across fields.",
        "complete",
        100,
        4,
        4,
        "",
        "Shipped (commit 77ae116).",
    ),
    (
        "F-012.3",
        "F-012",
        "Re-extraction on field edit",
        "When user edits a field, re-extract from the user's own prose if it looks like a paragraph.",
        "Pasting a paragraph into a single field offers extraction into multiple fields.",
        "not_started",
        0,
        8,
        0,
        "",
        "",
    ),
    # F-013 Feedback Moment 1
    (
        "F-013.1",
        "F-013",
        "Modal trigger after Draft Canvas",
        "Modal opens automatically the first time a user views their Draft Canvas.",
        "Modal appears once per user; dismissable; trigger not repeated.",
        "not_started",
        0,
        4,
        0,
        "",
        "",
    ),
    (
        "F-013.2",
        "F-013",
        "Feedback questions (per Launch Roadmap › Feedback Loop)",
        "Three short questions about whether the conversation revealed new territory.",
        "Questions match Launch Roadmap spec exactly.",
        "not_started",
        0,
        2,
        0,
        "",
        "",
    ),
    (
        "F-013.3",
        "F-013",
        "Submit to feedback sheet",
        "Response submits to the central feedback sheet (F-016).",
        "Row appears in feedback sheet within 10s of submit.",
        "not_started",
        0,
        4,
        0,
        "",
        "Depends on F-016.",
    ),
    # F-014 Feedback Moment 2
    (
        "F-014.1",
        "F-014",
        "Inline panel after Forensic Canvas",
        "Inline panel renders below the Forensic Canvas the first time it's viewed.",
        "Panel appears once per user; dismissable.",
        "not_started",
        0,
        4,
        0,
        "",
        "",
    ),
    (
        "F-014.2",
        "F-014",
        "'What will you do with this?' question with multi-select",
        "Multi-select including 'update existing assets' option (drives the asset-tracking decision gate).",
        "Selections recorded with timestamps; 'update existing assets' specifically tracked.",
        "not_started",
        0,
        4,
        0,
        "",
        "Critical for asset-tracking gate.",
    ),
    (
        "F-014.3",
        "F-014",
        "Submit to feedback sheet",
        "Response submits to the central feedback sheet.",
        "Row appears in feedback sheet within 10s of submit.",
        "not_started",
        0,
        4,
        0,
        "",
        "",
    ),
    # F-015 Feedback Moment 3
    (
        "F-015.1",
        "F-015",
        "Scheduled-send infrastructure",
        "Day-14 emails sent via Resend / Supabase cron; idempotent.",
        "Email sends exactly once 14 days after Forensic Canvas creation.",
        "not_started",
        0,
        4,
        0,
        "",
        "",
    ),
    (
        "F-015.2",
        "F-015",
        "Email content + reply parsing",
        "Email asks the question; replies parse into the feedback sheet.",
        "Replies flow into feedback sheet; failures alert Matthew.",
        "not_started",
        0,
        4,
        0,
        "",
        "",
    ),
    # F-016 Feedback response sheet integration
    (
        "F-016.1",
        "F-016",
        "Central feedback sheet schema",
        "One row per response; columns for moment, user_id, timestamp, payload JSON, parsed key fields.",
        "Schema documented; sheet (or DB view) exists.",
        "not_started",
        0,
        4,
        0,
        "",
        "",
    ),
    (
        "F-016.2",
        "F-016",
        "Write path from feedback moments",
        "All three feedback moments write to the central sheet.",
        "Test row from each moment appears in the sheet.",
        "not_started",
        0,
        4,
        0,
        "",
        "",
    ),
    # F-017 Error monitoring
    (
        "F-017.1",
        "F-017",
        "Sentry SDK installed",
        "Sentry installed for both client (Vite) and edge functions.",
        "Errors visible in Sentry dashboard within 1 min of occurrence.",
        "not_started",
        0,
        3,
        0,
        "",
        "",
    ),
    (
        "F-017.2",
        "F-017",
        "Daily review cadence documented",
        "Matthew has a documented daily review of Sentry during alpha.",
        "Cadence in tester onboarding doc; first review completed by alpha day 1.",
        "not_started",
        0,
        2,
        0,
        "",
        "",
    ),
    (
        "F-017.3",
        "F-017",
        "Source maps uploaded on build",
        "Production builds upload source maps to Sentry so stack traces resolve.",
        "Stack traces show original source file/line in Sentry UI.",
        "not_started",
        0,
        2,
        0,
        "",
        "",
    ),
    # F-018 Tester onboarding flow
    (
        "F-018.1",
        "F-018",
        "Invitation email template",
        "Warm, brief invitation framing 'first 10 people'; sets 7-day window + 30-min debrief expectation.",
        "Template approved by Trevor; sent through Resend.",
        "not_started",
        0,
        3,
        0,
        "",
        "",
    ),
    (
        "F-018.2",
        "F-018",
        "One-page Loom + written guide",
        "Short Loom recording + written one-pager covering: what we're testing, the three feedback moments, where to send bugs.",
        "Both assets live and linked from invitation email.",
        "not_started",
        0,
        3,
        0,
        "",
        "Matthew records the Loom.",
    ),
    (
        "F-018.3",
        "F-018",
        "In-app welcome screen",
        "Single-screen welcome acknowledging alpha state; names Trevor + framework.",
        "Welcome screen appears on first login; dismissable; no tutorial.",
        "not_started",
        0,
        2,
        0,
        "",
        "",
    ),
]


# Phase Gate items — mirror Alpha / Beta / GA Plan readiness gates.
# (phase, gate_name, gate_requirement, mapped_feature_ids, notes)
PHASE_GATES: list[tuple] = [
    (
        "Alpha",
        "Layer 1 conversation",
        "All 10 questions function end-to-end. Fallback responses written in Trevor's voice. Canvas sidebar populates in real time.",
        "F-001,F-002",
        "If the entry experience breaks, the tester never reaches anything else.",
    ),
    (
        "Alpha",
        "Draft Canvas",
        "Synthesis from Layer 1 produces a coherent canvas with DRAFT badge.",
        "F-003",
        "The Draft-to-Forensic differential is the product's core demonstration of value.",
    ),
    (
        "Alpha",
        "Avatar Builder",
        "All 4 sections accept input, persist, and surface AI assistance modes.",
        "F-004,F-005,F-006,F-007,F-009",
        "This is the product per Trevor's Respec.",
    ),
    (
        "Alpha",
        "Forensic Canvas",
        "All 28+ fields render. Trigger Moment displayed prominently as 'Strategic Entry Point'.",
        "F-008",
        "Trevor flagged Trigger Moment as the single most important field.",
    ),
    (
        "Alpha",
        "Manual edit lock",
        "Manually edited fields are flagged. AI never overwrites them silently.",
        "F-011",
        "Trust foundation. Existing PRD constraint.",
    ),
    (
        "Alpha",
        "Feedback capture",
        "Three feedback moments wired. Responses route to a single sheet Matthew can monitor.",
        "F-013,F-014,F-015,F-016",
        "No feedback infrastructure = no learning.",
    ),
    (
        "Alpha",
        "Error monitoring",
        "Sentry or equivalent active. Errors visible before testers report them.",
        "F-017",
        "",
    ),
    (
        "Alpha",
        "Tester onboarding doc",
        "One-page Loom + written guide. Sets expectation that this is alpha and frames feedback constructively.",
        "F-018",
        "",
    ),
    (
        "Alpha",
        "Architectural decision (Lovable vs existing)",
        "Decide explicitly which codebase is canonical for Gen 3. Don't half-port.",
        "F-052",
        "Risk Register: Critical. Two parallel codebases active >1 week triggers escalation.",
    ),
    # Beta
    (
        "Beta",
        "Alpha findings synthesised",
        "One-page findings doc. Top 3 product gaps named.",
        "F-016",
        "Synthesis is a Matthew + Trevor activity; tracked via feedback sheet readiness.",
    ),
    (
        "Beta",
        "Pay gate functional",
        "Stripe integrated. Both pricing options tested. Post-payment action menu wired.",
        "F-019",
        "Central conversion test of beta.",
    ),
    (
        "Beta",
        "Create mode v1",
        "At minimum: Amazon listing copy + Brand Voice Guide. Forensic canvas drives outputs.",
        "F-020,F-021",
        "Pay-gate previews need real outputs to anchor to.",
    ),
    (
        "Beta",
        "Coach mode v1",
        "Open conversation with full canvas as context. Document upload supported.",
        "F-022",
        "Required for the £29/mo subscription value prop.",
    ),
    (
        "Beta",
        "Export",
        "PDF + Word for both Draft and Forensic canvases.",
        "F-023",
        "",
    ),
    (
        "Beta",
        "Asset Inventory decision",
        "Built or formally deferred based on Alpha % flagging asset-update demand.",
        "F-028",
        "Default is NO. Decision lives in F-028 notes.",
    ),
    (
        "Beta",
        "Analytics + funnel tracking",
        "PostHog or Mixpanel events on every funnel step. Cohort analysis possible.",
        "F-024",
        "",
    ),
    (
        "Beta",
        "Support workflow",
        "Intercom (or email) inbox with documented SLA.",
        "F-025",
        "",
    ),
    (
        "Beta",
        "Brand Diagnostic integration",
        "Dynamic CTA, score handoff into Layer 1, refreshed visual identity.",
        "F-027",
        "",
    ),
    # GA
    (
        "GA",
        "My Brands dashboard",
        "Multi-brand support. Users can manage 2+ brands. Brand switcher works.",
        "F-031",
        "",
    ),
    (
        "GA",
        "Anonymous-to-authenticated migration",
        "Diagnostic completers without accounts have their data preserved when they sign up.",
        "F-032",
        "",
    ),
    (
        "GA",
        "Self-serve onboarding",
        "Tooltip-light onboarding tour. No human needed to start.",
        "F-033",
        "",
    ),
    (
        "GA",
        "Pricing finalised",
        "Pricing tested in beta. Tier structure final. Annual option live.",
        "F-034",
        "",
    ),
    (
        "GA",
        "Refund policy + ToS + Privacy",
        "Standard SaaS refund (14-day no-questions). Privacy policy. Terms of service.",
        "F-037",
        "",
    ),
    (
        "GA",
        "Status page",
        "Public uptime page. Incident communication channel.",
        "F-035",
        "",
    ),
    (
        "GA",
        "Knowledge base",
        "10–20 articles covering common questions from beta support.",
        "F-036",
        "",
    ),
    (
        "GA",
        "Pricing page",
        "Pricing page that converts cold traffic. Beta tested.",
        "F-034",
        "",
    ),
    (
        "GA",
        "Performance metrics tracking",
        "Per-avatar manual entry of ROAS, CTR, CPA, conversion rate, custom metrics.",
        "F-038",
        "OPEN QUESTION: Gen 3 spec does not address; existing V2 PRD scopes it. Treated here as GA P0 pending Matthew sign-off.",
    ),
]


# Risk register — mirrors Launch Roadmap › Risk Register sheet + adds current state.
# (risk_id, risk_name, phase, severity, current_state, trigger_threshold,
#  mitigation_owner, mitigation_status, notes)
RISKS: list[tuple] = [
    (
        "R-001",
        "Alpha testers don't complete the flow",
        "Alpha",
        "High",
        "monitoring",
        "<5 of 8 complete full flow by day 7",
        "Matthew",
        "Pre-flight QA checklist + daily monitoring + aggressive debrief follow-up planned",
        "",
    ),
    (
        "R-002",
        "Trevor's voice doesn't land for younger demographic",
        "Alpha",
        "Medium",
        "monitoring",
        "3+ testers flag 'sounds old' or 'corporate'",
        "Trevor",
        "Mixed-experience tester pool; tone question in feedback Moment 1",
        "",
    ),
    (
        "R-003",
        "Avatar Builder feels like a form not a conversation",
        "Alpha",
        "High",
        "monitoring",
        "Any tester says 'this is just a long form'",
        "Matthew",
        "F-009 AI assistance wired before alpha; F-010 progressive disclosure tested; question phrasing reviewed by Trevor",
        "Kills Trevor's core thesis if it materialises.",
    ),
    (
        "R-004",
        "Pay gate conversion too low",
        "Beta",
        "High",
        "monitoring",
        "<8% Draft viewers convert at 4-week mark",
        "Matthew",
        "Pay-gate previews use actual canvas data; pricing differentiated by job not features; A/B test framing",
        "",
    ),
    (
        "R-005",
        "Refund rate is high",
        "Beta",
        "Medium",
        "monitoring",
        ">10% refund rate",
        "Matthew",
        "Forensic canvas over-delivers vs pay-gate promise; 14-day refund window",
        "",
    ),
    (
        "R-006",
        "Asset tracking sucks Beta scope dry",
        "Beta",
        "Medium",
        "monitoring",
        "Builder spending >2 weeks on asset tracking in beta phase",
        "Matthew",
        "Decision gate at end of Alpha with hard threshold; v1 only if shipped",
        "",
    ),
    (
        "R-007",
        "ROI testimonials don't materialise",
        "Beta",
        "High",
        "monitoring",
        "<3 testimonials by week 4 of beta",
        "Matthew",
        "Testimonial harvest sequence baked in from day 1; results-focused user selection",
        "Kills the marketing engine if it materialises.",
    ),
    (
        "R-008",
        "GA acquisition CAC too high",
        "GA",
        "High",
        "monitoring",
        "Paid CAC payback >4 months after 30 days of spend",
        "Matthew",
        "Build organic foundation in beta; don't scale paid until CAC payback known",
        "",
    ),
    (
        "R-009",
        "Retention below unit economics floor",
        "GA",
        "Critical",
        "monitoring",
        "<50% retention at week 4",
        "Matthew",
        "Retention motion built before GA; quarterly Voice of Customer refresh designed in",
        "",
    ),
    (
        "R-010",
        "TikTok algorithm de-prioritises long-form B2B content",
        "Post-GA",
        "Medium",
        "monitoring",
        "Organic engagement <2% after 30 days of content",
        "Matthew",
        "Organic test before paid; cap spend at £500/wk in test phase",
        "",
    ),
    (
        "R-011",
        "Trevor's bandwidth limits content production",
        "All",
        "Medium",
        "monitoring",
        "Trevor unavailable >2 weeks consecutively",
        "Matthew",
        "Batch recording; async content review; develop a second voice as backup",
        "",
    ),
    (
        "R-012",
        "Lovable.dev vs existing codebase architectural collision",
        "Alpha",
        "Critical",
        "triggered",
        "Two parallel codebases active >1 week",
        "Matthew",
        "Decide explicitly at sprint start which is canonical (F-052)",
        "Status set to triggered because the decision is currently open and Gen 3 work is paused on this call.",
    ),
    (
        "R-013",
        "Competitor launches similar tool",
        "Beta",
        "Medium",
        "monitoring",
        "Competitor announcement or feature ship",
        "Matthew",
        "Speed; first-mover positioning on Forensic Avatar Builder; lean into Voice of Customer differentiator",
        "",
    ),
    (
        "R-014",
        "Trevor changes framework spec mid-build",
        "All",
        "High",
        "monitoring",
        "Any framework field change requested post-Alpha launch",
        "Matthew",
        "Lock the v2 Brief; require written sign-off on changes; version-tag the build",
        "",
    ),
]


# Brand Needs — concrete user-need rows that anchor the feature inventory to
# real brand use cases. Each leaf need maps to one or more feature ids so the
# Phase Gate / Summary sheets can be cross-referenced against actual demand.
# (need_id, brand, need, mapped_features, notes)
BRAND_NEEDS: list[tuple] = [
    (
        "N-001",
        "InfinityVault",
        "Create brand strategy doc using the Coach for the 'collector' avatar",
        "F-001,F-004,F-005,F-006,F-007,F-008,F-022,F-023",
        "End-to-end happy path: Layer 1 conversation → Avatar Builder (collector) → Forensic Canvas → Coach refinement → Export. Validates the core product loop with a real brand.",
    ),
    (
        "N-002",
        "InfinityVault",
        "Assess current set of images and videos",
        "F-054",
        "Top-level need under which the gap-identification and emotional-connection needs sit.",
    ),
    (
        "N-002.1",
        "InfinityVault",
        "Identify what's missing in the current image/video set",
        "F-054",
        "Gap analysis on visual assets — requires side-by-side comparison + tagging against the canvas.",
    ),
    (
        "N-002.1.1",
        "InfinityVault",
        "Surface opportunities to improve emotional connection with the customer",
        "F-006,F-007,F-054,F-055",
        "Connects Emotional Journey Mapping (Avatar Builder Sec 3) + Voice of Customer (Sec 4) to the asset gaps surfaced by F-054, with design briefs (F-055) closing the loop into action.",
    ),
]


# Change Log seed entry.
CHANGE_LOG: list[tuple] = [
    (
        "2026-05-13",
        "Matthew (via Claude Code)",
        "Initial workbook created",
        "Operational counterpart to Launch Roadmap; seeded from Gen 3 brief, V2 PRD, and Launch Roadmap workbook.",
    ),
    (
        TODAY.isoformat(),
        "Matthew (via Claude Code)",
        "Added F-053..F-056 + Brand Needs sheet",
        "Captured multi-angle, image comparison, design brief, and (post-GA) image generation features. Added Brand Needs sheet seeded with InfinityVault as the first validation case.",
    ),
]


# ---------------------------------------------------------------------------
# Workbook construction helpers
# ---------------------------------------------------------------------------


def set_column_widths(ws: Worksheet, widths: dict[str, int]) -> None:
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def write_header_row(ws: Worksheet, row: int, headers: list[str]) -> None:
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = CELL_BORDER
    ws.row_dimensions[row].height = 28


def write_title(ws: Worksheet, row: int, span: int, text: str) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = TITLE_FONT
    cell.fill = TITLE_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 32


def write_section(ws: Worksheet, row: int, span: int, text: str) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 22


def apply_status_conditional_formatting(
    ws: Worksheet, column_letter: str, first_row: int, last_row: int
) -> None:
    cell_range = f"{column_letter}{first_row}:{column_letter}{last_row}"
    for status, fill in STATUS_FILLS.items():
        rule = CellIsRule(
            operator="equal",
            formula=[f'"{status}"'],
            fill=fill,
        )
        ws.conditional_formatting.add(cell_range, rule)


def add_validation(
    ws: Worksheet,
    options: list[str],
    column_letter: str,
    first_row: int,
    last_row: int,
) -> None:
    formula = '"' + ",".join(options) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.error = "Value must match one of the allowed options."
    dv.errorTitle = "Invalid entry"
    ws.add_data_validation(dv)
    dv.add(f"{column_letter}{first_row}:{column_letter}{last_row}")


def style_body_cell(cell) -> None:
    cell.font = BODY_FONT
    cell.alignment = BODY_ALIGN
    cell.border = CELL_BORDER


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------


def build_features_sheet(wb: Workbook) -> Worksheet:
    ws = wb.create_sheet("Features")

    write_title(ws, 1, 15, "Features — implementation status by launch phase")
    headers = [
        "feature_id",
        "feature_name",
        "area",
        "description",
        "phase_required",
        "priority",
        "status",
        "percent_complete",
        "blocker",
        "owner",
        "depends_on",
        "effort_estimate_hours",
        "actual_hours",
        "notes",
        "last_updated",
    ]
    write_header_row(ws, 2, headers)

    first_data_row = 3
    for offset, feature in enumerate(FEATURES):
        row = first_data_row + offset
        (
            fid,
            name,
            area,
            description,
            phase,
            priority,
            status,
            percent,
            blocker,
            owner,
            depends_on,
            effort_h,
            actual_h,
            notes,
        ) = feature
        values = [
            fid,
            name,
            area,
            description,
            phase,
            priority,
            status,
            percent,
            blocker,
            owner,
            depends_on,
            effort_h,
            actual_h,
            notes,
            TODAY,
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            style_body_cell(cell)
        # last_updated as date
        ws.cell(row=row, column=15).number_format = "yyyy-mm-dd"
        ws.cell(row=row, column=8).number_format = "0"
        ws.row_dimensions[row].height = 42

    last_data_row = first_data_row + len(FEATURES) - 1

    # Data validation on enum columns
    add_validation(ws, AREA_OPTIONS, "C", first_data_row, last_data_row + 200)
    add_validation(ws, PHASE_OPTIONS, "E", first_data_row, last_data_row + 200)
    add_validation(ws, PRIORITY_OPTIONS, "F", first_data_row, last_data_row + 200)
    add_validation(ws, STATUS_OPTIONS, "G", first_data_row, last_data_row + 200)

    # Conditional formatting on status column
    apply_status_conditional_formatting(ws, "G", first_data_row, last_data_row + 200)

    # Freeze panes below header
    ws.freeze_panes = "C3"

    set_column_widths(
        ws,
        {
            "A": 10,
            "B": 38,
            "C": 16,
            "D": 60,
            "E": 14,
            "F": 9,
            "G": 14,
            "H": 9,
            "I": 32,
            "J": 12,
            "K": 18,
            "L": 10,
            "M": 10,
            "N": 60,
            "O": 12,
        },
    )

    return ws


def build_subfeatures_sheet(wb: Workbook) -> Worksheet:
    ws = wb.create_sheet("Sub-features")
    write_title(ws, 1, 11, "Sub-features — shippable units per feature")
    headers = [
        "subfeature_id",
        "parent_feature_id",
        "subfeature_name",
        "description",
        "acceptance_criteria",
        "status",
        "percent_complete",
        "effort_estimate_hours",
        "actual_hours",
        "blocker",
        "notes",
    ]
    write_header_row(ws, 2, headers)

    first_data_row = 3
    for offset, sub in enumerate(SUBFEATURES):
        row = first_data_row + offset
        (
            sid,
            parent,
            name,
            description,
            criteria,
            status,
            percent,
            effort_h,
            actual_h,
            blocker,
            notes,
        ) = sub
        values = [
            sid,
            parent,
            name,
            description,
            criteria,
            status,
            percent,
            effort_h,
            actual_h,
            blocker,
            notes,
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            style_body_cell(cell)
        ws.cell(row=row, column=7).number_format = "0"
        ws.row_dimensions[row].height = 42

    last_data_row = first_data_row + len(SUBFEATURES) - 1

    add_validation(ws, STATUS_OPTIONS, "F", first_data_row, last_data_row + 200)
    apply_status_conditional_formatting(ws, "F", first_data_row, last_data_row + 200)

    ws.freeze_panes = "C3"

    set_column_widths(
        ws,
        {
            "A": 12,
            "B": 12,
            "C": 38,
            "D": 50,
            "E": 50,
            "F": 14,
            "G": 9,
            "H": 10,
            "I": 10,
            "J": 28,
            "K": 38,
        },
    )

    return ws


def build_summary_sheet(wb: Workbook, n_features: int) -> Worksheet:
    """Summary built first (positioned at index 0); references Features sheet
    by name, so it works regardless of build order."""
    ws = wb.create_sheet("Summary", 0)

    write_title(ws, 1, 8, "Summary — feature readiness by launch phase")
    ws.cell(
        row=2,
        column=1,
        value=(
            "Live rollup from the Features sheet. Status column derives on-track / at-risk / blocked from the underlying data."
        ),
    ).font = BODY_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)

    headers = [
        "phase",
        "features_required",
        "complete",
        "in_progress",
        "blocked",
        "not_started",
        "percent_complete",
        "status",
    ]
    write_header_row(ws, 4, headers)

    features_range = f"'Features'!$E$3:$E${3 + n_features - 1}"
    status_range = f"'Features'!$G$3:$G${3 + n_features - 1}"
    percent_range = f"'Features'!$H$3:$H${3 + n_features - 1}"

    phase_row_map: dict[str, int] = {}
    for offset, phase in enumerate(PHASE_OPTIONS):
        row = 5 + offset
        phase_row_map[phase] = row
        ws.cell(row=row, column=1, value=phase)
        ws.cell(
            row=row,
            column=2,
            value=f'=COUNTIF({features_range},"{phase}")',
        )
        ws.cell(
            row=row,
            column=3,
            value=(
                f'=COUNTIFS({features_range},"{phase}",{status_range},"complete")'
            ),
        )
        ws.cell(
            row=row,
            column=4,
            value=(
                f'=COUNTIFS({features_range},"{phase}",{status_range},"in_progress")'
                f'+COUNTIFS({features_range},"{phase}",{status_range},"in_review")'
            ),
        )
        ws.cell(
            row=row,
            column=5,
            value=(
                f'=COUNTIFS({features_range},"{phase}",{status_range},"blocked")'
            ),
        )
        ws.cell(
            row=row,
            column=6,
            value=(
                f'=COUNTIFS({features_range},"{phase}",{status_range},"not_started")'
            ),
        )
        ws.cell(
            row=row,
            column=7,
            value=(
                f'=IFERROR(ROUND(SUMIFS({percent_range},{features_range},"{phase}")/(B{row}*100),3),0)'
            ),
        )
        ws.cell(
            row=row,
            column=8,
            value=(
                f'=IF(E{row}>0,"BLOCKED",'
                f'IF(G{row}>=0.9,"ON-TRACK",'
                f'IF(G{row}>=0.5,"AT-RISK",'
                f'IF(B{row}=0,"N/A","AT-RISK"))))'
            ),
        )
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            style_body_cell(cell)
        ws.cell(row=row, column=7).number_format = "0%"

    # Earliest blocker — read row by row from Features sheet
    last_phase_row = 5 + len(PHASE_OPTIONS) - 1
    blocker_row = last_phase_row + 2
    write_section(ws, blocker_row, 8, "Earliest blocker (first feature with status = blocked)")
    blocker_id_cell = ws.cell(
        row=blocker_row + 1,
        column=1,
        value=(
            f'=IFERROR(INDEX(\'Features\'!A3:A{3 + n_features - 1},'
            f'MATCH("blocked",{status_range},0)),"None")'
        ),
    )
    blocker_name_cell = ws.cell(
        row=blocker_row + 1,
        column=2,
        value=(
            f'=IFERROR(INDEX(\'Features\'!B3:B{3 + n_features - 1},'
            f'MATCH("blocked",{status_range},0)),"None")'
        ),
    )
    blocker_phase_cell = ws.cell(
        row=blocker_row + 1,
        column=3,
        value=(
            f'=IFERROR(INDEX(\'Features\'!E3:E{3 + n_features - 1},'
            f'MATCH("blocked",{status_range},0)),"")'
        ),
    )
    blocker_reason_cell = ws.cell(
        row=blocker_row + 1,
        column=4,
        value=(
            f'=IFERROR(INDEX(\'Features\'!I3:I{3 + n_features - 1},'
            f'MATCH("blocked",{status_range},0)),"")'
        ),
    )
    ws.merge_cells(
        start_row=blocker_row + 1,
        start_column=4,
        end_row=blocker_row + 1,
        end_column=8,
    )
    for cell in [blocker_id_cell, blocker_name_cell, blocker_phase_cell, blocker_reason_cell]:
        style_body_cell(cell)
    ws.row_dimensions[blocker_row + 1].height = 28

    # Conditional formatting on Status column (column H)
    cell_range = f"H5:H{last_phase_row}"
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="equal", formula=['"BLOCKED"'], fill=STATUS_FILLS["blocked"]),
    )
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="equal", formula=['"AT-RISK"'], fill=STATUS_FILLS["in_progress"]),
    )
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="equal", formula=['"ON-TRACK"'], fill=STATUS_FILLS["complete"]),
    )

    ws.freeze_panes = "A5"

    set_column_widths(
        ws,
        {
            "A": 14,
            "B": 24,
            "C": 14,
            "D": 14,
            "E": 12,
            "F": 14,
            "G": 18,
            "H": 14,
        },
    )

    return ws


def build_phase_gate_sheet(wb: Workbook, n_features: int) -> Worksheet:
    ws = wb.create_sheet("Phase Gate")
    write_title(ws, 1, 6, "Phase Gate — readiness gates by launch phase")
    ws.cell(
        row=2,
        column=1,
        value=(
            "Mirrors the readiness gates from the Launch Roadmap. gate_status = MET when every mapped feature is complete."
        ),
    ).font = BODY_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)

    headers = [
        "phase",
        "gate_name",
        "gate_requirement",
        "mapped_features",
        "gate_status",
        "notes",
    ]
    write_header_row(ws, 4, headers)

    feature_id_range = f"'Features'!$A$3:$A${3 + n_features - 1}"
    feature_status_range = f"'Features'!$G$3:$G${3 + n_features - 1}"

    first_data_row = 5
    for offset, gate in enumerate(PHASE_GATES):
        row = first_data_row + offset
        phase, name, requirement, mapped, notes = gate
        ws.cell(row=row, column=1, value=phase)
        ws.cell(row=row, column=2, value=name)
        ws.cell(row=row, column=3, value=requirement)
        ws.cell(row=row, column=4, value=mapped)

        # gate_status: count completed mapped features vs total mapped count.
        # Build a SUMPRODUCT expression: number of mapped feature ids whose
        # status is "complete".
        ids = [i.strip() for i in mapped.split(",") if i.strip()]
        if not ids:
            status_formula = '"N/A"'
        else:
            total = len(ids)
            terms = []
            for fid in ids:
                # COUNTIFS returns 1 if the matched feature has status complete.
                terms.append(
                    f'COUNTIFS({feature_id_range},"{fid}",{feature_status_range},"complete")'
                )
            complete_sum = "(" + "+".join(terms) + ")"
            blocked_terms = [
                f'COUNTIFS({feature_id_range},"{fid}",{feature_status_range},"blocked")'
                for fid in ids
            ]
            blocked_sum = "(" + "+".join(blocked_terms) + ")"
            status_formula = (
                f'=IF({blocked_sum}>0,"BLOCKED — see mapped features",'
                f'IF({complete_sum}={total},"MET",'
                f'"OPEN — "&({total}-{complete_sum})&" of {total} remaining"))'
            )
        ws.cell(row=row, column=5, value=status_formula)
        ws.cell(row=row, column=6, value=notes)
        for col in range(1, 7):
            style_body_cell(ws.cell(row=row, column=col))
        ws.row_dimensions[row].height = 42

    last_data_row = first_data_row + len(PHASE_GATES) - 1

    add_validation(ws, PHASE_OPTIONS, "A", first_data_row, last_data_row + 100)

    # Conditional formatting on gate_status (column E) — match a few sentinels.
    # FormulaRule references the first cell of the range; Excel/openpyxl handles
    # the relative shift across the rest of the range.
    cell_range = f"E{first_data_row}:E{last_data_row}"
    ws.conditional_formatting.add(
        cell_range,
        FormulaRule(formula=[f'E{first_data_row}="MET"'], fill=STATUS_FILLS["complete"]),
    )
    ws.conditional_formatting.add(
        cell_range,
        FormulaRule(
            formula=[f'ISNUMBER(SEARCH("BLOCKED",E{first_data_row}))'],
            fill=STATUS_FILLS["blocked"],
        ),
    )
    ws.conditional_formatting.add(
        cell_range,
        FormulaRule(
            formula=[f'ISNUMBER(SEARCH("OPEN",E{first_data_row}))'],
            fill=STATUS_FILLS["in_progress"],
        ),
    )

    ws.freeze_panes = "A5"

    set_column_widths(
        ws,
        {
            "A": 10,
            "B": 36,
            "C": 60,
            "D": 28,
            "E": 38,
            "F": 50,
        },
    )

    return ws


def build_risks_sheet(wb: Workbook) -> Worksheet:
    ws = wb.create_sheet("Risks")
    write_title(ws, 1, 9, "Risks — register with current state")
    headers = [
        "risk_id",
        "risk_name",
        "phase_exposed",
        "severity",
        "current_state",
        "trigger_threshold",
        "mitigation_owner",
        "mitigation_status",
        "notes",
    ]
    write_header_row(ws, 3, headers)

    first_data_row = 4
    for offset, risk in enumerate(RISKS):
        row = first_data_row + offset
        for col_idx, value in enumerate(risk, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            style_body_cell(cell)
        ws.row_dimensions[row].height = 42

    last_data_row = first_data_row + len(RISKS) - 1

    add_validation(
        ws, PHASE_OPTIONS + ["All"], "C", first_data_row, last_data_row + 100
    )
    add_validation(ws, SEVERITY_OPTIONS, "D", first_data_row, last_data_row + 100)
    add_validation(ws, RISK_STATE_OPTIONS, "E", first_data_row, last_data_row + 100)

    # Conditional formatting on current_state
    cell_range = f"E{first_data_row}:E{last_data_row + 100}"
    state_fills = {
        "monitoring": STATUS_FILLS["not_started"],
        "triggered": STATUS_FILLS["blocked"],
        "mitigated": STATUS_FILLS["in_progress"],
        "resolved": STATUS_FILLS["complete"],
    }
    for state, fill in state_fills.items():
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(operator="equal", formula=[f'"{state}"'], fill=fill),
        )

    # Conditional formatting on severity
    sev_range = f"D{first_data_row}:D{last_data_row + 100}"
    sev_fills = {
        "Critical": PatternFill("solid", fgColor="FFFF7575"),
        "High": STATUS_FILLS["blocked"],
        "Medium": STATUS_FILLS["in_progress"],
        "Low": STATUS_FILLS["not_started"],
    }
    for sev, fill in sev_fills.items():
        ws.conditional_formatting.add(
            sev_range,
            CellIsRule(operator="equal", formula=[f'"{sev}"'], fill=fill),
        )

    ws.freeze_panes = "B4"

    set_column_widths(
        ws,
        {
            "A": 8,
            "B": 40,
            "C": 14,
            "D": 10,
            "E": 14,
            "F": 38,
            "G": 14,
            "H": 50,
            "I": 38,
        },
    )

    return ws


def build_brand_needs_sheet(wb: Workbook, n_features: int) -> Worksheet:
    ws = wb.create_sheet("Brand Needs")
    write_title(ws, 1, 6, "Brand Needs — concrete user needs mapped to features")
    ws.cell(
        row=2,
        column=1,
        value=(
            "Anchors the feature inventory to real brand use cases. Each leaf need maps to one or more feature ids. "
            "fulfilment_status is derived: MET if all mapped features complete, BLOCKED if any are blocked, otherwise OPEN with count."
        ),
    ).font = BODY_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)

    headers = [
        "need_id",
        "brand",
        "need",
        "mapped_features",
        "fulfilment_status",
        "notes",
    ]
    write_header_row(ws, 4, headers)

    feature_id_range = f"'Features'!$A$3:$A${3 + n_features - 1}"
    feature_status_range = f"'Features'!$G$3:$G${3 + n_features - 1}"

    first_data_row = 5
    for offset, need in enumerate(BRAND_NEEDS):
        row = first_data_row + offset
        need_id, brand, text, mapped, notes = need
        ws.cell(row=row, column=1, value=need_id)
        ws.cell(row=row, column=2, value=brand)
        ws.cell(row=row, column=3, value=text)
        ws.cell(row=row, column=4, value=mapped)

        ids = [i.strip() for i in mapped.split(",") if i.strip()]
        if not ids:
            status_formula = '"N/A"'
        else:
            total = len(ids)
            complete_terms = [
                f'COUNTIFS({feature_id_range},"{fid}",{feature_status_range},"complete")'
                for fid in ids
            ]
            blocked_terms = [
                f'COUNTIFS({feature_id_range},"{fid}",{feature_status_range},"blocked")'
                for fid in ids
            ]
            complete_sum = "(" + "+".join(complete_terms) + ")"
            blocked_sum = "(" + "+".join(blocked_terms) + ")"
            status_formula = (
                f'=IF({blocked_sum}>0,"BLOCKED — see mapped features",'
                f'IF({complete_sum}={total},"MET",'
                f'"OPEN — "&({total}-{complete_sum})&" of {total} remaining"))'
            )
        ws.cell(row=row, column=5, value=status_formula)
        ws.cell(row=row, column=6, value=notes)
        for col in range(1, 7):
            style_body_cell(ws.cell(row=row, column=col))
        ws.row_dimensions[row].height = 42

    last_data_row = first_data_row + len(BRAND_NEEDS) - 1

    # Conditional formatting on fulfilment_status (column E)
    cell_range = f"E{first_data_row}:E{last_data_row}"
    ws.conditional_formatting.add(
        cell_range,
        FormulaRule(formula=[f'E{first_data_row}="MET"'], fill=STATUS_FILLS["complete"]),
    )
    ws.conditional_formatting.add(
        cell_range,
        FormulaRule(
            formula=[f'ISNUMBER(SEARCH("BLOCKED",E{first_data_row}))'],
            fill=STATUS_FILLS["blocked"],
        ),
    )
    ws.conditional_formatting.add(
        cell_range,
        FormulaRule(
            formula=[f'ISNUMBER(SEARCH("OPEN",E{first_data_row}))'],
            fill=STATUS_FILLS["in_progress"],
        ),
    )

    ws.freeze_panes = "C5"

    set_column_widths(
        ws,
        {
            "A": 12,
            "B": 18,
            "C": 60,
            "D": 30,
            "E": 36,
            "F": 50,
        },
    )

    return ws


def build_change_log_sheet(wb: Workbook) -> Worksheet:
    ws = wb.create_sheet("Change Log")
    write_title(ws, 1, 4, "Change Log — append-only")
    headers = ["date", "who", "what_changed", "reason"]
    write_header_row(ws, 3, headers)

    first_data_row = 4
    for offset, entry in enumerate(CHANGE_LOG):
        row = first_data_row + offset
        for col_idx, value in enumerate(entry, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            style_body_cell(cell)
        ws.row_dimensions[row].height = 28

    ws.freeze_panes = "A4"

    set_column_widths(ws, {"A": 14, "B": 28, "C": 50, "D": 60})

    return ws


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> None:
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    # Drop the default sheet
    default = wb.active
    wb.remove(default)

    build_features_sheet(wb)
    build_subfeatures_sheet(wb)
    n_features = len(FEATURES)
    build_summary_sheet(wb, n_features)
    build_phase_gate_sheet(wb, n_features)
    build_risks_sheet(wb)
    build_brand_needs_sheet(wb, n_features)
    build_change_log_sheet(wb)

    # Reorder: Summary first, then Features inventory, then derived views.
    order = [
        "Summary",
        "Features",
        "Sub-features",
        "Phase Gate",
        "Brand Needs",
        "Risks",
        "Change Log",
    ]
    wb._sheets = [wb[name] for name in order]

    wb.save(OUTPUT_PATH)
    print(f"Wrote {OUTPUT_PATH}")
    print(f"Features: {len(FEATURES)} | Sub-features: {len(SUBFEATURES)} | "
          f"Phase gates: {len(PHASE_GATES)} | Brand needs: {len(BRAND_NEEDS)} | "
          f"Risks: {len(RISKS)}")


if __name__ == "__main__":
    main()
