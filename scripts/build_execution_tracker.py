#!/usr/bin/env python3
"""Build the IDEA Brand Coach Gen 3 Execution & Time Tracker workbook.

Triangulates effort from three sources, each flagged so logged vs. estimated
time is never conflated:
  - logged    : real Clockify entries (docs/data/clockify_export.csv)
  - meeting   : Fathom meeting durations for meetings NOT already in Clockify
                (docs/data/fathom_meetings.json)
  - estimated : Claude Code JSONL session active-time (~/.claude/projects/*idea-brand-coach*)

Output: docs/execution_tracker.xlsx (6 sheets, SUMIFS/COUNTIFS rollups).

This is the execution + effort view at a strategic-objective grain. It links to
the feature-level feature_status_tracker.xlsx by feature ID but does not
duplicate its detail.

Usage:
    python3 scripts/build_execution_tracker.py
    python3 scripts/build_execution_tracker.py --clockify /path/to/export.csv --out docs/execution_tracker.xlsx
"""
from __future__ import annotations

import argparse
import csv
import glob
import json
import os
import re
from datetime import datetime, date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------- #
# Paths
# --------------------------------------------------------------------------- #
REPO_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
DEFAULT_CLOCKIFY = os.path.join(REPO_ROOT, "docs", "data", "clockify_export.csv")
DEFAULT_FATHOM = os.path.join(REPO_ROOT, "docs", "data", "fathom_meetings.json")
DEFAULT_OUT = os.path.join(REPO_ROOT, "docs", "execution_tracker.xlsx")
JSONL_GLOB = os.path.expanduser("~/.claude/projects/*idea-brand-coach*")

CLOCKIFY_PROJECTS = ("Brand Coach Beta", "IDEA Brand Coach Marketing")
IDLE_CAP_SECONDS = 15 * 60  # cap inter-event gaps when estimating active time

# --------------------------------------------------------------------------- #
# Strategic objectives (the tracked rows). Rolls up feature IDs from the
# feature_status_tracker; rollup_pct is a placeholder to be reconciled against
# that workbook's 2026-05-16 audit.
# --------------------------------------------------------------------------- #
OBJECTIVES = [
    ("O1", "Diagnostic & Trust Gap™", "F-001, F-002",
     "Layer 1 conversation: 10-question diagnostic + sidebar population/follow-up",
     "Alpha", "in_progress", 0.30),
    ("O2", "Draft Canvas Synthesis", "F-003",
     "Synthesize Layer 1 answers into a draft brand canvas",
     "Alpha", "in_progress", 0.42),
    ("O3", "Customer Insight Engine / Avatar Builder", "F-004, F-005, F-006",
     "Avatar Builder sections 1-4, AI assistance, confidence scoring & edit lock",
     "Alpha/Beta", "in_progress", 0.45),
    ("O4", "The Signature & Coach Mode v2", "F-009",
     "Stage 5 retention output: brand voice guide, Amazon listing copy",
     "Beta", "in_progress", 0.45),
    ("O5", "Export & Deliverables", "F-007",
     "PDF / Word / Markdown export of brand strategy",
     "Beta", "in_progress", 0.72),
    ("O6", "Monetization / Pay Gate", "F-008",
     "Stripe pay gate before premium outputs",
     "GA", "not_started", 0.0),
    ("O7", "Launch & Strategy", "—",
     "Alpha->Beta->GA gates, ideabrandcoach.com site, TikTok-conditional, Trevor sync, roadmap/context upkeep",
     "All", "in_progress", None),
]
OBJ_IDS = [o[0] for o in OBJECTIVES]

# Keyword -> objective rules, evaluated in order; first match wins.
# Tuned to the descriptions present in the Clockify export.
KEYWORD_RULES = [
    (("extraction field", "field review", "field sync", "field visibility"), "O3"),
    (("avatar", "insight engine", "psychographic"), "O3"),
    (("export", "pdf", "docx", "brand book"), "O5"),
    (("stripe", "pay gate", "paywall", "payment"), "O6"),
    (("listing copy", "brand voice", "signature", "coach mode"), "O4"),
    (("diagnostic", "trust gap", "chat", "agent sdk", "broken text", "hardcoded text"), "O1"),
    (("canvas", "synthesis"), "O2"),
    # O7 catch-alls: collaboration, planning, marketing, ops
    (("trevor", "whatsapp", "reply", "message", "meet", "mock up", "mockup",
      "feedback", "workbook", "spreadsheet", "todos", "plan", "remotion",
      "marketing video", "case study", "funnel", "obsidian", "autoresearch",
      "lovable", "ramp up", "review"), "O7"),
]


def map_description_to_objective(desc: str) -> tuple[str, str]:
    """Return (objective_id, note). note flags unmatched rows for review."""
    d = (desc or "").lower()
    for keywords, obj in KEYWORD_RULES:
        if any(k in d for k in keywords):
            return obj, ""
    return "O7", "review: unmatched -> default O7"


# --------------------------------------------------------------------------- #
# Source 1: Clockify
# --------------------------------------------------------------------------- #
def parse_duration_decimal(row: dict) -> float:
    val = (row.get("Duration (decimal)") or "").strip()
    try:
        return float(val)
    except ValueError:
        return 0.0


def parse_clockify_date(row: dict) -> date | None:
    raw = (row.get("Start Date") or "").strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


def load_clockify(path: str) -> list[dict]:
    """Return time-log rows from Clockify, source='logged'."""
    rows = []
    with open(path, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            if row.get("Project", "").strip() not in CLOCKIFY_PROJECTS:
                continue
            hours = parse_duration_decimal(row)
            if hours <= 0:
                continue
            desc = (row.get("Description") or "").strip()
            obj, note = map_description_to_objective(desc)
            rows.append({
                "date": parse_clockify_date(row),
                "source": "logged",
                "objective_id": obj,
                "description": desc,
                "hours": round(hours, 3),
                "note": note,
            })
    return rows


# --------------------------------------------------------------------------- #
# Source 2: Fathom meetings (only unlogged ones add hours)
# --------------------------------------------------------------------------- #
def load_fathom(path: str) -> tuple[list[dict], list[dict]]:
    """Return (time_log_rows, meeting_sheet_rows).

    Only meetings with in_clockify=False produce time-log 'meeting' rows;
    the rest are already counted in Clockify and shown for reference only.
    """
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    time_rows, meeting_rows = [], []
    for m in data.get("meetings", []):
        d = datetime.strptime(m["date"], "%Y-%m-%d").date()
        meeting_rows.append({
            "date": d,
            "title": m["title"],
            "recording_id": m["recording_id"],
            "url": m["url"],
            "duration_h": m["duration_h"],
            "in_clockify": "Y" if m["in_clockify"] else "N",
            "objective_id": m.get("objective_id", "O7"),
        })
        if not m["in_clockify"]:
            time_rows.append({
                "date": d,
                "source": "meeting",
                "objective_id": m.get("objective_id", "O7"),
                "description": f"Meeting: {m['title']}",
                "hours": round(float(m["duration_h"]), 3),
                "note": m.get("note", ""),
            })
    return time_rows, meeting_rows


# --------------------------------------------------------------------------- #
# Source 3: JSONL session active-time (estimated)
# --------------------------------------------------------------------------- #
def _ts(obj: dict) -> datetime | None:
    t = obj.get("timestamp")
    if not t:
        return None
    try:
        return datetime.fromisoformat(t.replace("Z", "+00:00"))
    except ValueError:
        return None


def load_sessions(glob_root: str) -> tuple[list[dict], list[dict]]:
    """Return (time_log_rows, session_sheet_rows) from top-level JSONL sessions."""
    time_rows, session_rows = [], []
    for d in glob.glob(glob_root):
        worktree = os.path.basename(d).split("--claude-worktrees-")[-1]
        for fn in glob.glob(os.path.join(d, "*.jsonl")):  # top-level only, skip subagents/
            times = []
            for line in open(fn, encoding="utf-8"):
                line = line.strip()
                if not line:
                    continue
                try:
                    o = json.loads(line)
                except json.JSONDecodeError:
                    continue
                t = _ts(o)
                if t:
                    times.append(t)
            if len(times) < 2:
                continue
            times.sort()
            active = sum(
                min((times[i] - times[i - 1]).total_seconds(), IDLE_CAP_SECONDS)
                for i in range(1, len(times))
            )
            active_h = round(active / 3600.0, 3)
            if active_h <= 0:
                continue
            sess_date = times[0].date()
            sid = os.path.splitext(os.path.basename(fn))[0]
            session_rows.append({
                "date": sess_date,
                "session_id": sid[:8],
                "worktree": worktree,
                "event_count": len(times),
                "active_h": active_h,
                "objective_id": "O7",
            })
            time_rows.append({
                "date": sess_date,
                "source": "estimated",
                "objective_id": "O7",
                "description": f"Agent session {sid[:8]} ({worktree}, {len(times)} events)",
                "hours": active_h,
                "note": "estimated from JSONL active time (15-min idle cap)",
            })
    return time_rows, session_rows


# --------------------------------------------------------------------------- #
# Workbook styling helpers
# --------------------------------------------------------------------------- #
HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=14)
SUBTLE = Font(italic=True, color="6B7280")
THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
SOURCE_FILL = {
    "logged": PatternFill("solid", fgColor="DCFCE7"),
    "meeting": PatternFill("solid", fgColor="DBEAFE"),
    "estimated": PatternFill("solid", fgColor="FEF3C7"),
}


def write_header(ws, headers, row=1):
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def autosize(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# --------------------------------------------------------------------------- #
# Sheet builders
# --------------------------------------------------------------------------- #
def build_time_log(wb, time_rows):
    ws = wb.create_sheet("Time Log")
    headers = ["Date", "Source", "Objective", "Description", "Hours", "Note"]
    write_header(ws, headers)
    time_rows.sort(key=lambda r: (r["date"] or date.min, r["source"]))
    for i, r in enumerate(time_rows, start=2):
        ws.cell(row=i, column=1, value=r["date"]).number_format = "yyyy-mm-dd"
        sc = ws.cell(row=i, column=2, value=r["source"])
        sc.fill = SOURCE_FILL.get(r["source"], PatternFill())
        ws.cell(row=i, column=3, value=r["objective_id"])
        ws.cell(row=i, column=4, value=r["description"])
        ws.cell(row=i, column=5, value=r["hours"]).number_format = "0.00"
        ws.cell(row=i, column=6, value=r["note"]).font = SUBTLE
        for c in range(1, 7):
            ws.cell(row=i, column=c).border = BORDER
    autosize(ws, {"A": 12, "B": 11, "C": 11, "D": 60, "E": 9, "F": 42})
    return ws, len(time_rows)


def build_objectives(wb, n_timelog):
    ws = wb.create_sheet("Objectives")
    headers = ["ID", "Objective", "Linked Features", "Description", "Phase",
               "Status", "Rollup %", "Hours Logged", "Hours Meeting",
               "Hours Estimated", "Hours Total"]
    write_header(ws, headers)
    last = n_timelog + 1  # Time Log data rows: 2..last
    rng_obj = f"'Time Log'.$C$2:$C${last}"  # placeholder, fixed below
    for i, (oid, name, feats, desc, phase, status, pct) in enumerate(OBJECTIVES, start=2):
        ws.cell(row=i, column=1, value=oid)
        ws.cell(row=i, column=2, value=name)
        ws.cell(row=i, column=3, value=feats)
        ws.cell(row=i, column=4, value=desc)
        ws.cell(row=i, column=5, value=phase)
        ws.cell(row=i, column=6, value=status)
        pc = ws.cell(row=i, column=7, value=(pct if pct is not None else "n/a"))
        if pct is not None:
            pc.number_format = "0%"
        # SUMIFS over Time Log: by objective (+ source for the split columns)
        oc = f"$C$2:$C${last}"   # objective col
        hc = f"$E$2:$E${last}"   # hours col
        sc = f"$B$2:$B${last}"   # source col
        ws.cell(row=i, column=8,
                value=f"=SUMIFS('Time Log'!{hc},'Time Log'!{oc},$A{i},'Time Log'!{sc},\"logged\")").number_format = "0.00"
        ws.cell(row=i, column=9,
                value=f"=SUMIFS('Time Log'!{hc},'Time Log'!{oc},$A{i},'Time Log'!{sc},\"meeting\")").number_format = "0.00"
        ws.cell(row=i, column=10,
                value=f"=SUMIFS('Time Log'!{hc},'Time Log'!{oc},$A{i},'Time Log'!{sc},\"estimated\")").number_format = "0.00"
        ws.cell(row=i, column=11, value=f"=SUM(H{i}:J{i})").number_format = "0.00"
        for c in range(1, 12):
            ws.cell(row=i, column=c).border = BORDER
    # Totals row
    tot = len(OBJECTIVES) + 2
    ws.cell(row=tot, column=2, value="TOTAL").font = Font(bold=True)
    for c in (8, 9, 10, 11):
        col = get_column_letter(c)
        cell = ws.cell(row=tot, column=c, value=f"=SUM({col}2:{col}{tot-1})")
        cell.number_format = "0.00"
        cell.font = Font(bold=True)
    autosize(ws, {"A": 6, "B": 36, "C": 18, "D": 56, "E": 11, "F": 13,
                  "G": 9, "H": 13, "I": 13, "J": 15, "K": 12})
    return ws


def build_summary(wb, totals):
    ws = wb.create_sheet("Summary")
    ws.cell(row=1, column=1, value="IDEA Brand Coach Gen 3 — Execution & Time Tracker").font = TITLE_FONT
    ws.cell(row=2, column=1,
            value=f"Generated {date.today().isoformat()} · execution + effort view at strategic-objective grain · "
                  f"links to feature_status_tracker.xlsx by feature ID").font = SUBTLE

    # Time-by-source block
    ws.cell(row=4, column=1, value="Time invested by source").font = Font(bold=True)
    write_header(ws, ["Source", "Hours", "Share"], row=5)
    grand = sum(totals.values()) or 1.0
    order = [("logged", "Logged (Clockify)"), ("meeting", "Meetings (Fathom, unlogged)"),
             ("estimated", "Estimated (agent sessions)")]
    r = 6
    for key, label in order:
        ws.cell(row=r, column=1, value=label).fill = SOURCE_FILL.get(key, PatternFill())
        ws.cell(row=r, column=2, value=round(totals.get(key, 0.0), 2)).number_format = "0.00"
        ws.cell(row=r, column=3, value=totals.get(key, 0.0) / grand).number_format = "0%"
        for c in range(1, 4):
            ws.cell(row=r, column=c).border = BORDER
        r += 1
    ws.cell(row=r, column=1, value="GRAND TOTAL").font = Font(bold=True)
    ws.cell(row=r, column=2, value=round(grand, 2)).number_format = "0.00"
    ws.cell(row=r, column=2).font = Font(bold=True)
    for c in range(1, 4):
        ws.cell(row=r, column=c).border = BORDER

    # Per-objective rollup referencing Objectives sheet
    start = r + 2
    ws.cell(row=start - 1, column=1, value="Per-objective rollup").font = Font(bold=True)
    write_header(ws, ["Objective", "Status", "Rollup %", "Total Hours"], row=start)
    for i in range(len(OBJECTIVES)):
        rr = start + 1 + i
        orow = 2 + i  # Objectives data rows
        ws.cell(row=rr, column=1, value=f"=Objectives!B{orow}")
        ws.cell(row=rr, column=2, value=f"=Objectives!F{orow}")
        ws.cell(row=rr, column=3, value=f"=Objectives!G{orow}")
        ws.cell(row=rr, column=4, value=f"=Objectives!K{orow}").number_format = "0.00"
        for c in range(1, 5):
            ws.cell(row=rr, column=c).border = BORDER
    autosize(ws, {"A": 40, "B": 14, "C": 10, "D": 12})
    # Make Summary the first sheet
    wb.move_sheet("Summary", -(len(wb.sheetnames) - 1))
    return ws


def build_meetings(wb, meeting_rows):
    ws = wb.create_sheet("Meetings")
    headers = ["Date", "Title", "Recording ID", "URL", "Duration (h)",
               "In Clockify?", "Objective"]
    write_header(ws, headers)
    meeting_rows.sort(key=lambda r: r["date"], reverse=True)
    for i, m in enumerate(meeting_rows, start=2):
        ws.cell(row=i, column=1, value=m["date"]).number_format = "yyyy-mm-dd"
        ws.cell(row=i, column=2, value=m["title"])
        ws.cell(row=i, column=3, value=m["recording_id"])
        ws.cell(row=i, column=4, value=m["url"])
        ws.cell(row=i, column=5, value=m["duration_h"]).number_format = "0.00"
        ic = ws.cell(row=i, column=6, value=m["in_clockify"])
        ic.fill = PatternFill("solid", fgColor="DCFCE7" if m["in_clockify"] == "Y" else "FEE2E2")
        ws.cell(row=i, column=7, value=m["objective_id"])
        for c in range(1, 8):
            ws.cell(row=i, column=c).border = BORDER
    ws.cell(row=len(meeting_rows) + 3, column=2,
            value="N = meeting time NOT in Clockify (added as 'meeting' hours). "
                  "Y = already counted in Clockify (reference only).").font = SUBTLE
    autosize(ws, {"A": 12, "B": 48, "C": 13, "D": 42, "E": 12, "F": 13, "G": 10})
    return ws


def build_sessions(wb, session_rows):
    ws = wb.create_sheet("Sessions")
    headers = ["Date", "Session ID", "Worktree", "Events", "Active (h, est.)", "Objective"]
    write_header(ws, headers)
    session_rows.sort(key=lambda r: r["date"], reverse=True)
    for i, s in enumerate(session_rows, start=2):
        ws.cell(row=i, column=1, value=s["date"]).number_format = "yyyy-mm-dd"
        ws.cell(row=i, column=2, value=s["session_id"])
        ws.cell(row=i, column=3, value=s["worktree"])
        ws.cell(row=i, column=4, value=s["event_count"])
        ws.cell(row=i, column=5, value=s["active_h"]).number_format = "0.00"
        ws.cell(row=i, column=6, value=s["objective_id"])
        for c in range(1, 7):
            ws.cell(row=i, column=c).border = BORDER
    n = len(session_rows)
    ws.cell(row=n + 2, column=4, value="TOTAL").font = Font(bold=True)
    ws.cell(row=n + 2, column=5, value=f"=SUM(E2:E{n+1})").number_format = "0.00"
    ws.cell(row=n + 2, column=5).font = Font(bold=True)
    ws.cell(row=n + 3, column=2,
            value="Active time = sum of inter-event gaps, capped at 15 min/gap. "
                  "Under-counts: subagent work is not in top-level sessions.").font = SUBTLE
    autosize(ws, {"A": 12, "B": 12, "C": 22, "D": 9, "E": 16, "F": 10})
    return ws


def build_changelog(wb):
    ws = wb.create_sheet("Change Log")
    write_header(ws, ["Date", "Version", "Change"])
    ws.cell(row=2, column=1, value=date.today()).number_format = "yyyy-mm-dd"
    ws.cell(row=2, column=2, value="v1.0")
    ws.cell(row=2, column=3,
            value="Initial build: 7 strategic objectives; time triangulated from "
                  "Clockify (logged) + Fathom (meeting) + JSONL (estimated).")
    autosize(ws, {"A": 12, "B": 10, "C": 80})
    return ws


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #
def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--clockify", default=DEFAULT_CLOCKIFY)
    ap.add_argument("--fathom", default=DEFAULT_FATHOM)
    ap.add_argument("--out", default=DEFAULT_OUT)
    args = ap.parse_args()

    clock_rows = load_clockify(args.clockify)
    fathom_time, meeting_rows = load_fathom(args.fathom)
    sess_time, session_rows = load_sessions(JSONL_GLOB)

    time_rows = clock_rows + fathom_time + sess_time
    totals = {"logged": 0.0, "meeting": 0.0, "estimated": 0.0}
    for r in time_rows:
        totals[r["source"]] = totals.get(r["source"], 0.0) + r["hours"]

    wb = Workbook()
    wb.remove(wb.active)  # drop default sheet
    _, n_timelog = build_time_log(wb, time_rows)
    build_objectives(wb, n_timelog)
    build_meetings(wb, meeting_rows)
    build_sessions(wb, session_rows)
    build_changelog(wb)
    build_summary(wb, totals)  # moves itself to front

    os.makedirs(os.path.dirname(args.out), exist_ok=True)
    wb.save(args.out)

    print(f"Wrote {args.out}")
    print(f"  Time Log rows : {n_timelog}")
    print(f"  Logged    : {totals['logged']:.2f} h  ({len(clock_rows)} Clockify entries)")
    print(f"  Meeting   : {totals['meeting']:.2f} h  ({len(fathom_time)} unlogged meetings)")
    print(f"  Estimated : {totals['estimated']:.2f} h  ({len(session_rows)} agent sessions)")
    print(f"  GRAND     : {sum(totals.values()):.2f} h")


if __name__ == "__main__":
    main()
