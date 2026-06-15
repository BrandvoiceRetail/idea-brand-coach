#!/usr/bin/env python3
"""
Structural consistency comparator for the gold-workbook output engine (Phase 7 Agent B).

Compares two runs of the SAME workbook for STRUCTURAL stability while tolerating PROSE
variation (the LLM legs are non-deterministic). Used by scripts/rehearsal-consistency.ts.

Usage:
    python3 xlsx_structural_diff.py <run1_A> <run2_A> <run1_B> <run2_B>

Consistency contract (a structural drift -> exit 1):
  - identical sheet set + order;
  - identical first-column STRUCTURAL labels per sheet (banners, section headers, table
    header rows, tier tokens, fixed enum labels) — these are the skeleton, not the prose;
  - same number of vocabulary clusters / signature options / audit-IDEA rows (±1);
  - same matrix tiering (T1/T2/T3 data-row counts, exact) and same rollout phase count (exact).

PROSE cells (the free-text body of a row: a cluster's description, a canvas value, a
brief example) are NOT compared — only the structural labels in column A and the counts.
The structural labels are stable strings emitted verbatim by the deterministic assemblers
(assembleWorkbookA.ts / assembleWorkbookB.ts), so any divergence there is real drift.
"""
import sys
import openpyxl


# First-column labels that are STRUCTURAL (assembler-emitted skeleton). We compare the
# full ordered list of column-A labels per sheet, which captures banners, section headers,
# table-header rows, and the fixed enum rows (Insight/Distinctiveness/…, T1/T2/T3, Phase…).
# Free-text data lands in OTHER columns or is a model sentence; the column-A label of a
# data row in these gold sheets is itself a fixed token (cluster name is column A on the
# avatar sheet, but cluster NAMES are deterministic category labels, not model prose).
def column_a_labels(ws):
    labels = []
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        if not any(c is not None and str(c).strip() != "" for c in row):
            continue
        first = row[0]
        labels.append(str(first).strip() if first is not None else "")
    return labels


def sheet_dims(ws):
    rows = [
        r for r in ws.iter_rows(values_only=True)
        if r and any(c is not None and str(c).strip() != "" for c in r)
    ]
    maxcol = 0
    for r in rows:
        last = 0
        for i, c in enumerate(r):
            if c is not None and str(c).strip() != "":
                last = i + 1
        maxcol = max(maxcol, last)
    return len(rows), maxcol


def load(path):
    return openpyxl.load_workbook(path, data_only=True)


def diff_workbook(label, p1, p2, drift):
    """Compare workbook p2 (run 2) against p1 (run 1). Append drift strings to `drift`."""
    print(f"\n[{label}] run1={p1}")
    print(f"[{label}] run2={p2}")
    wb1, wb2 = load(p1), load(p2)

    names1 = [ws.title for ws in wb1.worksheets]
    names2 = [ws.title for ws in wb2.worksheets]
    if names1 != names2:
        drift.append(f"{label}: sheet set/order differs: run1={names1} run2={names2}")
        print(f"  DRIFT sheet set: run1={names1} run2={names2}")
        return  # cannot align sheets meaningfully

    print(f"  sheets ({len(names1)}): {names1}")
    for ws1, ws2 in zip(wb1.worksheets, wb2.worksheets):
        labels1 = column_a_labels(ws1)
        labels2 = column_a_labels(ws2)
        n1, c1 = sheet_dims(ws1)
        n2, c2 = sheet_dims(ws2)
        same = labels1 == labels2
        print(
            f"  - '{ws1.title}': rows {n1}->{n2}, cols {c1}->{c2}, "
            f"colA-labels {'MATCH' if same else 'DIFFER'} ({len(labels1)} vs {len(labels2)})"
        )
        if c1 != c2:
            drift.append(f"{label}/'{ws1.title}': column count drift {c1} -> {c2}")
        if not same:
            # show the first divergence for diagnosis
            first_div = None
            for i, (a, b) in enumerate(zip(labels1, labels2)):
                if a != b:
                    first_div = (i, a, b)
                    break
            if first_div is None and len(labels1) != len(labels2):
                drift.append(
                    f"{label}/'{ws1.title}': colA-label COUNT drift {len(labels1)} -> {len(labels2)}"
                )
                print(f"      label-count drift: {len(labels1)} -> {len(labels2)}")
            else:
                idx, a, b = first_div
                drift.append(
                    f"{label}/'{ws1.title}': colA-label drift at row {idx}: '{a}' -> '{b}'"
                )
                print(f"      first label drift @row {idx}: '{a}' -> '{b}'")


def count_label_prefix(labels, prefix):
    return sum(1 for x in labels if x.strip() == prefix or x.strip().startswith(prefix))


def semantic_counts(label, p1, p2, drift):
    """The ±1 cluster/option/audit-row counts and exact tiering/phase counts."""
    wb1, wb2 = load(p1), load(p2)
    sheets1 = {ws.title: column_a_labels(ws) for ws in wb1.worksheets}
    sheets2 = {ws.title: column_a_labels(ws) for ws in wb2.worksheets}

    def tol1(name, getter, what):
        if name not in sheets1 or name not in sheets2:
            return
        a, b = getter(sheets1[name]), getter(sheets2[name])
        ok = abs(a - b) <= 1
        print(f"  {what}: run1={a} run2={b} ({'OK ±1' if ok else 'DRIFT >1'})")
        if not ok:
            drift.append(f"{label}: {what} drifted >1: {a} -> {b}")

    def exact(name, getter, what):
        if name not in sheets1 or name not in sheets2:
            return
        a, b = getter(sheets1[name]), getter(sheets2[name])
        ok = a == b
        print(f"  {what}: run1={a} run2={b} ({'OK exact' if ok else 'DRIFT'})")
        if not ok:
            drift.append(f"{label}: {what} differs (must be exact): {a} -> {b}")

    if label == "Workbook A":
        # Avatar sheet: count cluster rows (between the 'Cluster' header and the
        # 'Stage 5' banner) and signature option rows.
        def cluster_count(labels):
            try:
                start = labels.index("Cluster") + 1
            except ValueError:
                return 0
            n = 0
            for x in labels[start:]:
                if x.startswith("Stage 5") or x == "Option":
                    break
                n += 1
            return n

        def option_count(labels):
            return count_label_prefix(labels, "Option ")

        def audit_row_count(labels):
            # Audit×IDEA rows: column-A labels after 'Audit investment' header.
            try:
                start = labels.index("Audit investment") + 1
            except ValueError:
                return 0
            return len(labels) - start

        tol1("4. Avatar 2.0 (IV)", cluster_count, "vocabulary clusters")
        tol1("4. Avatar 2.0 (IV)", option_count, "signature options")
        tol1("7. Audit × IDEA", audit_row_count, "audit×IDEA investment rows")

    if label == "Workbook B":
        def tier_count(tier):
            return lambda labels: sum(1 for x in labels if x.strip() == tier)

        def phase_count(labels):
            return count_label_prefix(labels, "Phase ")

        exact("Investment Matrix", tier_count("T1"), "T1 matrix rows")
        exact("Investment Matrix", tier_count("T2"), "T2 matrix rows")
        exact("Investment Matrix", tier_count("T3"), "T3 matrix rows")
        exact("Recommended Phasing", phase_count, "rollout phases")


def main():
    if len(sys.argv) != 5:
        print("usage: xlsx_structural_diff.py <run1_A> <run2_A> <run1_B> <run2_B>", file=sys.stderr)
        sys.exit(2)
    run1_a, run2_a, run1_b, run2_b = sys.argv[1:5]
    drift = []

    diff_workbook("Workbook A", run1_a, run2_a, drift)
    semantic_counts("Workbook A", run1_a, run2_a, drift)
    diff_workbook("Workbook B", run1_b, run2_b, drift)
    semantic_counts("Workbook B", run1_b, run2_b, drift)

    print("\n=== STRUCTURAL DIFF RESULT ===")
    if drift:
        print(f"{len(drift)} structural drift(s):")
        for d in drift:
            print(f"  - {d}")
        sys.exit(1)
    print("No structural drift — run-2 is structurally consistent with run-1.")
    sys.exit(0)


if __name__ == "__main__":
    main()
